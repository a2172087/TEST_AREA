import warnings
warnings.filterwarnings('ignore', category=DeprecationWarning)
warnings.filterwarnings('ignore', message='Workbook contains no default style')

import sys
import os
import re
import subprocess
import pyodbc
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import plotly.graph_objects as go
import plotly.offline
import tempfile
import webbrowser
import json
from datetime import datetime
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import socket
import shutil
from flask import Flask, jsonify, request, send_from_directory, redirect
from tkinter import Tk, filedialog

# 修復 stdout/stderr None 問題
if sys.stdout is None:
    sys.stdout = open(os.devnull, 'w')
if sys.stderr is None:
    sys.stderr = open(os.devnull, 'w')

# 修復高 DPI 螢幕模糊問題
try:
    from ctypes import windll
    windll.shcore.SetProcessDpiAwareness(1)
except:
    pass

# 系統設定
username = os.environ.get('USERNAME', 'Unknown')

# SQL Server 連線資訊
json_path = r"M:\BI_Database\Apps\Database\Apps_Database\O_All\SQL_Server\SQL_Server_Info_User_BI.json"
with open(json_path, 'r') as file:
    sql_connection_info = json.load(file)

SQL_SERVER_INFO = {
    "server": sql_connection_info["server"],
    "database": sql_connection_info["database"], 
    "username": sql_connection_info["username"],
    "password": sql_connection_info["password"],
    "apps_log_table": sql_connection_info["apps_log_table"]  
}

# 網路資源路徑
NETWORK_ASSETS = {
    "font_awesome_base": r"M:\BI_Database\Apps\Database\Apps_Database\Q_All\Wafer_Map_Stack_Analysis\Font_Awesome",
    "google_fonts_base": r"M:\BI_Database\Apps\Database\Apps_Database\Q_All\Wafer_Map_Stack_Analysis\Google_Fonts",
}

# Flask 應用程式初始化
app = Flask(__name__)

# 全域變數
current_worker = None
last_activity_time = time.time()
ACTIVITY_TIMEOUT = 30 * 60
selected_folder_path = None
xlsx_file_list = []
analysis_file_data = None
anomaly_weights = {}
selected_process_type = 'CP'

class FileData:
    """儲存所有檔案資料"""
    def __init__(self):
        self.standard_file_name = ""
        self.cust_prod_id = ""
        self.cust_prod_group = ""
        self.prod_group = ""
        self.wafer_notch = ""
        
        self.wafer_map_x_range = 0
        self.wafer_map_y_range = 0
        self.wafer_map_die_count = 0
        self.wafer_map_die_point = []
        
        self.min_x = None
        self.max_x = None
        self.min_y = None
        self.max_y = None
        
        self.die_anomaly_data = {}
        self.valid_colors = ['000000', 'C6E2FF']
        self.total_files_processed = 0
        self.total_sheets_processed = 0
        
        self.anomaly_code_names = {}
        self.anomaly_code_colors = {}


class AnalysisWorker(threading.Thread):
    """多執行緒資料分析處理"""
    
    def __init__(self, xlsx_files, anomaly_weights, progress_callback=None, 
                complete_callback=None, error_callback=None):
        super().__init__()
        self.xlsx_files = xlsx_files
        self.anomaly_weights = anomaly_weights
        self.file_data = None
        self.should_stop = False
        self.data_lock = threading.RLock()
        self.progress_lock = threading.RLock()
        self.reference_prod_group = None
        self.prod_group_inconsistent = False
        
        self.progress_callback = progress_callback
        self.complete_callback = complete_callback
        self.error_callback = error_callback
        
        self.current_progress = 0
        self.status = 'idle'
        self.error_message = ''
        self.status_message = 'Initializing...'
        
        self.total_files = len(xlsx_files)
        self.files_completed = 0
        self.current_file_name = ''
        self.current_file_total_sheets = 0
        self.current_file_processed_sheets = 0

    def update_progress(self, value, status_msg=''):
        """更新進度百分比和狀態訊息"""
        with self.progress_lock:
            self.current_progress = value
            if status_msg:
                self.status_message = status_msg
            if self.progress_callback:
                self.progress_callback(value, status_msg)

    def stop(self):
        """停止執行緒"""
        self.should_stop = True

    def find_cell_in_column_a(self, df, search_value, file_name=""):
        """在 DataFrame 的 A 欄中搜尋特定值"""
        try:
            column_a = df.iloc[:, 0]
            
            for idx, cell_value in enumerate(column_a):
                if pd.notna(cell_value):
                    cell_str = str(cell_value).strip()
                    if cell_str.lower() == search_value.lower():
                        return idx
            
            raise ValueError(f"Cannot find '{search_value}' in column A" + 
                            (f" of file {file_name}" if file_name else ""))
            
        except ValueError:
            raise
        except Exception as e:
            raise Exception(f"Error searching for '{search_value}' in column A: {str(e)}")

    def find_cell_in_worksheet_column_a(self, ws, search_value, file_name=""):
        """在工作表的 A 欄中搜尋特定值"""
        try:
            for row in ws.iter_rows(min_col=1, max_col=1):
                cell = row[0]
                if cell.value:
                    cell_str = str(cell.value).strip()
                    if cell_str.lower() == search_value.lower():
                        return cell.row
            
            raise ValueError(f"Cannot find '{search_value}' in column A" + 
                            (f" of file {file_name}" if file_name else ""))
            
        except ValueError:
            raise
        except Exception as e:
            raise Exception(f"Error searching for '{search_value}' in worksheet column A: {str(e)}")

    def get_wafer_map_range(self, df):
        """確定 Wafer Map 的 X 和 Y 範圍"""
        try:
            x_max = 0
            row_index = 0
            col_start = 3
            
            first_row = df.iloc[row_index, col_start:]
            for idx, cell_value in enumerate(first_row):
                if pd.notna(cell_value):
                    try:
                        x_max = int(cell_value)
                    except:
                        pass
                else:
                    break
            
            self.file_data.wafer_map_x_range = x_max
            
            y_max = 0
            col_index = 2
            row_start = 1
            
            third_col = df.iloc[row_start:, col_index]
            for idx, cell_value in enumerate(third_col):
                if pd.notna(cell_value):
                    try:
                        y_max = int(cell_value)
                    except:
                        pass
                else:
                    break
            
            self.file_data.wafer_map_y_range = y_max
            
        except Exception as e:
            print(f"Error getting wafer map range: {str(e)}")

    def analyze_wafer_map(self, file_path):
        """分析 Wafer Map 找出所有 Die 位置"""
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb[wb.sheetnames[0]]
            
            self.file_data.wafer_map_die_count = 0
            self.file_data.wafer_map_die_point = []
            
            start_col = 4
            start_row = 2
            
            end_col = start_col + self.file_data.wafer_map_x_range
            end_row = start_row + self.file_data.wafer_map_y_range
            
            y = 0
            for row in ws.iter_rows(min_row=start_row, max_row=end_row, 
                                min_col=start_col, max_col=end_col):
                x = 0
                for cell in row:
                    has_die = False
                    
                    if cell.fill and cell.fill.patternType:
                        if cell.fill.patternType != 'none' and cell.fill.patternType is not None:
                            if cell.fill.fgColor and hasattr(cell.fill.fgColor, 'rgb'):
                                bg_color = str(cell.fill.fgColor.rgb)
                                if bg_color and bg_color not in ['FFFFFF', '00FFFFFF', 'FFFFFFFF', None, 'None']:
                                    has_die = True
                    
                    if has_die:
                        self.file_data.wafer_map_die_count += 1
                        self.file_data.wafer_map_die_point.append((x, y))
                    
                    x += 1
                y += 1
            
            wb.close()
            
        except Exception as e:
            print(f"Error analyzing wafer map: {str(e)}")

    def get_cell_background_color(self, cell):
        """取得儲存格背景顏色的 RGB 值"""
        try:
            if cell.fill:
                if hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                    if hasattr(cell.fill.fgColor, 'rgb'):
                        color = str(cell.fill.fgColor.rgb)
                        if len(color) >= 6:
                            return color[-6:]
                    elif hasattr(cell.fill.fgColor, 'value'):
                        color = str(cell.fill.fgColor.value)
                        if len(color) >= 6:
                            return color[-6:]
            return None
        except Exception as e:
            print(f"Error getting cell background color: {str(e)}")
            return None

    def get_specific_anomaly_code_info(self, ws, code):
        """搜尋特定異常代碼的名稱和顏色資訊"""
        try:
            if code in self.file_data.anomaly_code_names:
                return
            
            col_index = 1
            start_row = 13
            
            for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, 
                                min_col=col_index, max_col=col_index):
                cell = row[0]
                cell_value = cell.value
                
                if cell_value and isinstance(cell_value, str):
                    if ' : ' in cell_value:
                        parts = cell_value.split(' : ')
                        if len(parts) >= 2:
                            cell_code = parts[0].strip()
                            
                            if cell_code == code:
                                name = parts[1].strip()
                                self.file_data.anomaly_code_names[code] = name
                                
                                bg_color = self.get_cell_background_color(cell)
                                if bg_color:
                                    self.file_data.anomaly_code_colors[code] = bg_color
                                
                                return
                                
        except Exception as e:
            print(f"Error getting anomaly code info: {str(e)}")

    def process_single_file(self, file_path):
        """處理單一檔案的異常分析"""
        file_name = os.path.basename(file_path)
        local_anomaly_data = {}
        local_code_names = {}
        local_code_colors = {}
        sheets_processed = 0
        error_info = None
        
        try:
            wb = load_workbook(file_path, data_only=True)
            
            total_sheets_in_file = len([s for s in wb.sheetnames if s != 'Wafer Summary'])
            
            with self.progress_lock:
                self.current_file_name = file_name
                self.current_file_total_sheets = total_sheets_in_file
                self.current_file_processed_sheets = 0
            
            print(f"Processing file: {file_name} (Total sheets: {total_sheets_in_file})")
            
            ws_first = wb[wb.sheetnames[0]]
            
            # 檢查 Product Group
            current_prod_group = None
            for row in ws_first.iter_rows(min_col=1, max_col=2):
                if row[0].value and str(row[0].value).strip().lower() == 'prod group':
                    current_prod_group = str(row[1].value) if row[1].value else None
                    break
            
            if not current_prod_group:
                wb.close()
                error_info = {
                    'type': 'MissingRequiredField',
                    'file': file_name,
                    'message': f"Cannot find 'Prod Group' in column A or its value is empty."
                }
                return {
                    'file_name': file_name,
                    'anomaly_data': local_anomaly_data,
                    'code_names': local_code_names,
                    'code_colors': local_code_colors,
                    'sheets_processed': sheets_processed,
                    'error': error_info
                }
            
            # 驗證 Product Group 一致性
            with self.data_lock:
                if self.reference_prod_group is None:
                    self.reference_prod_group = current_prod_group
                    print(f"Reference Product Group set: {current_prod_group}")
                else:
                    if current_prod_group != self.reference_prod_group:
                        print(f"Product Group mismatch found in {file_name}")
                        print(f"  Expected: {self.reference_prod_group}")
                        print(f"  Found: {current_prod_group}")
                        
                        self.prod_group_inconsistent = True
                        
                        error_info = {
                            'type': 'ProductGroupMismatch',
                            'file': file_name,
                            'message': f'Product Group mismatch detected!\n\nFile: {file_name}\nExpected: {self.reference_prod_group}\nFound: {current_prod_group}\n\nAll files must have the same Product Group!',
                            'expected': self.reference_prod_group,
                            'found': current_prod_group
                        }
                        
                        wb.close()
                        return {
                            'file_name': file_name,
                            'anomaly_data': local_anomaly_data,
                            'code_names': local_code_names,
                            'code_colors': local_code_colors,
                            'sheets_processed': sheets_processed,
                            'error': error_info
                        }
            
            # 尋找 Total Defect Count 位置
            try:
                defect_count_row = self.find_cell_in_worksheet_column_a(
                    ws_first, 
                    'Total Defect Count', 
                    file_name
                )
                print(f"  Found 'Total Defect Count' at row {defect_count_row} in {file_name}")
                
            except ValueError as e:
                wb.close()
                error_info = {
                    'type': 'MissingRequiredField',
                    'file': file_name,
                    'message': f"Cannot find 'Total Defect Count' in column A.\nThis field is required for anomaly code analysis."
                }
                return {
                    'file_name': file_name,
                    'anomaly_data': local_anomaly_data,
                    'code_names': local_code_names,
                    'code_colors': local_code_colors,
                    'sheets_processed': sheets_processed,
                    'error': error_info
                }
            
            # 處理每個工作表
            for sheet_name in wb.sheetnames:
                if sheet_name == 'Wafer Summary':
                    continue
                
                ws = wb[sheet_name]
                
                start_col = 4
                start_row = 2
                
                min_x = self.file_data.min_x
                max_x = self.file_data.max_x
                min_y = self.file_data.min_y
                max_y = self.file_data.max_y
                
                # 快取儲存格資料
                cell_cache = {}
                for row_idx, row in enumerate(ws.iter_rows(
                    min_row=start_row + min_y, 
                    max_row=start_row + max_y,
                    min_col=start_col + min_x, 
                    max_col=start_col + max_x)):
                    for col_idx, cell in enumerate(row):
                        actual_x = min_x + col_idx
                        actual_y = min_y + row_idx
                        cell_cache[(actual_x, actual_y)] = cell
                
                # 檢查每個 Die 位置的異常
                for (x, y) in self.file_data.wafer_map_die_point:
                    if (x, y) in cell_cache:
                        cell = cell_cache[(x, y)]
                        cell_value = cell.value
                        
                        if cell_value is not None:
                            bg_color = self.get_cell_background_color(cell)
                            
                            if bg_color and bg_color.upper() not in self.file_data.valid_colors:
                                if isinstance(cell_value, (int, float)):
                                    value_str = str(int(cell_value))
                                else:
                                    value_str = str(cell_value)
                                
                                # 搜尋異常代碼資訊
                                if value_str not in local_code_names and value_str not in self.file_data.anomaly_code_names:
                                    search_end_row = min(defect_count_row + 100, ws.max_row)
                                    
                                    for row_info in ws.iter_rows(min_row=defect_count_row, max_row=search_end_row, min_col=1, max_col=1):
                                        info_cell = row_info[0]
                                        info_value = info_cell.value
                                        
                                        if info_value and isinstance(info_value, str) and ' : ' in info_value:
                                            parts = info_value.split(' : ')
                                            if len(parts) >= 2:
                                                cell_code = parts[0].strip()
                                                
                                                try:
                                                    cell_code = str(int(float(cell_code)))
                                                except:
                                                    pass
                                                
                                                if cell_code == value_str:
                                                    name = parts[1].strip()
                                                    local_code_names[value_str] = name
                                                    
                                                    info_bg_color = self.get_cell_background_color(info_cell)
                                                    if info_bg_color:
                                                        local_code_colors[value_str] = info_bg_color
                                                    
                                                    break
                                
                                if (x, y) not in local_anomaly_data:
                                    local_anomaly_data[(x, y)] = {}
                                
                                if value_str in local_anomaly_data[(x, y)]:
                                    local_anomaly_data[(x, y)][value_str] += 1
                                else:
                                    local_anomaly_data[(x, y)][value_str] = 1
                
                sheets_processed += 1
                
                # 更新進度
                with self.progress_lock:
                    self.current_file_processed_sheets = sheets_processed
                    
                    files_progress = self.files_completed / self.total_files
                    current_file_progress = sheets_processed / total_sheets_in_file / self.total_files
                    overall_progress = 20 + int((files_progress + current_file_progress) * 70)
                    
                    overall_progress = min(overall_progress, 90)
                    
                    status_msg = f"Processing [{self.files_completed + 1}/{self.total_files}] {file_name} - Sheet {sheets_processed}/{total_sheets_in_file}"
                    self.update_progress(overall_progress, status_msg)
                
                print(f"  Completed sheet: {sheet_name} ({sheets_processed}/{total_sheets_in_file})")
            
            wb.close()
            
            print(f"File processing completed: {file_name} (Total sheets processed: {sheets_processed})")
            
            return {
                'file_name': file_name,
                'anomaly_data': local_anomaly_data,
                'code_names': local_code_names,
                'code_colors': local_code_colors,
                'sheets_processed': sheets_processed,
                'error': error_info
            }
            
        except PermissionError:
            error_info = {
                'type': 'PermissionError',
                'file': file_name,
                'message': f'Cannot access the file. Please ensure:\n1. The file is not open in Excel\n2. You have read permission for the file'
            }
            return {
                'file_name': file_name,
                'anomaly_data': local_anomaly_data,
                'code_names': local_code_names,
                'code_colors': local_code_colors,
                'sheets_processed': sheets_processed,
                'error': error_info
            }
        except Exception as e:
            error_info = {
                'type': 'ProcessingError',
                'file': file_name,
                'message': str(e)
            }
            return {
                'file_name': file_name,
                'anomaly_data': local_anomaly_data,
                'code_names': local_code_names,
                'code_colors': local_code_colors,
                'sheets_processed': sheets_processed,
                'error': error_info
            }

    def analyze_all_files_anomalies(self):
        """使用多執行緒分析所有檔案的異常狀況"""
        try:
            total_files = len(self.xlsx_files)
            files_processed = 0
            errors_encountered = []
            
            with ThreadPoolExecutor(max_workers=6) as executor:
                futures = {}
                for file_idx, file_path in enumerate(self.xlsx_files):
                    if self.should_stop:
                        executor.shutdown(wait=False)
                        return
                        
                    file_name = os.path.basename(file_path)
                    print(f"Submitting file {file_idx + 1}/{total_files}: {file_name}")
                    future = executor.submit(self.process_single_file, file_path)
                    futures[future] = file_path
                
                for future in as_completed(futures):
                    if self.should_stop:
                        executor.shutdown(wait=False)
                        return
                        
                    try:
                        result = future.result()
                        
                        if result['error']:
                            errors_encountered.append(result['error'])
                            
                            if result['error']['type'] == 'MissingRequiredField':
                                self.should_stop = True
                                executor.shutdown(wait=False)
                                
                                self.error_message = result['error']['message']
                                self.status = 'error'
                                if self.error_callback:
                                    self.error_callback('Missing Required Field', result['error']['message'])
                                return
                            
                            elif result['error']['type'] == 'ProductGroupMismatch':
                                self.should_stop = True
                                executor.shutdown(wait=False)
                                
                                self.error_message = result['error']['message']
                                self.status = 'error'
                                if self.error_callback:
                                    self.error_callback('Product Group Mismatch', result['error']['message'])
                                return
                            
                            elif result['error']['type'] == 'PermissionError':
                                executor.shutdown(wait=False)
                                self.error_message = f"File: {result['error']['file']}\n\n{result['error']['message']}"
                                self.status = 'error'
                                if self.error_callback:
                                    self.error_callback('File Access Error', self.error_message)
                                return
                        else:
                            # 合併處理結果
                            with self.data_lock:
                                for coord, anomalies in result['anomaly_data'].items():
                                    if coord not in self.file_data.die_anomaly_data:
                                        self.file_data.die_anomaly_data[coord] = {}
                                    
                                    for value_str, count in anomalies.items():
                                        if value_str in self.file_data.die_anomaly_data[coord]:
                                            self.file_data.die_anomaly_data[coord][value_str] += count
                                        else:
                                            self.file_data.die_anomaly_data[coord][value_str] = count
                                
                                for code, name in result['code_names'].items():
                                    if code not in self.file_data.anomaly_code_names:
                                        self.file_data.anomaly_code_names[code] = name
                                
                                for code, color in result['code_colors'].items():
                                    if code not in self.file_data.anomaly_code_colors:
                                        self.file_data.anomaly_code_colors[code] = color
                                
                                self.file_data.total_sheets_processed += result['sheets_processed']
                            
                            files_processed += 1
                            self.file_data.total_files_processed = files_processed
                            
                            with self.progress_lock:
                                self.files_completed = files_processed
                                progress_value = 20 + int((files_processed / total_files) * 70)
                                self.update_progress(progress_value, 
                                                f"Completed {files_processed}/{total_files} files")
                        
                    except Exception as e:
                        error_msg = f"Error processing future result: {str(e)}"
                        print(error_msg)
                        
                        import traceback
                        print(f"Traceback:\n{traceback.format_exc()}")
                        
                        self.should_stop = True
                        self.error_message = error_msg
                        self.status = 'error'
                        
                        if self.error_callback:
                            self.error_callback('Processing Error', error_msg)
                        
                        executor.shutdown(wait=False)
                        return
            
            # 顯示分析結果摘要
            print("\nAnomaly Analysis Complete:")
            print(f"  Total files processed: {self.file_data.total_files_processed}")
            print(f"  Total sheets processed: {self.file_data.total_sheets_processed}")
            
            if self.file_data.anomaly_code_names:
                print("\nAnomaly Code Mapping:")
                for code in sorted(self.file_data.anomaly_code_names.keys()):
                    name = self.file_data.anomaly_code_names[code]
                    color = self.file_data.anomaly_code_colors.get(code, 'Unknown')
                    print(f"  Code '{code}': {name} (Color: {color})")
            
            if self.file_data.die_anomaly_data:
                print(f"\nAnomaly Summary (Total {len(self.file_data.die_anomaly_data)} coordinates with anomalies):")
                for idx, (coord, values) in enumerate(sorted(self.file_data.die_anomaly_data.items())):
                    if idx >= 5:
                        print(f"  ... and {len(self.file_data.die_anomaly_data) - 5} more coordinates")
                        break
                    print(f"  Coordinate {coord}:")
                    for value, count in sorted(values.items(), key=lambda x: x[1], reverse=True):
                        name = self.file_data.anomaly_code_names.get(value, '')
                        if name:
                            print(f"    - Value '{value}' ({name}) appeared {count} times")
                        else:
                            print(f"    - Value '{value}' appeared {count} times")
            else:
                print("\nNo anomalies found in any files.")
                
        except Exception as e:
            self.error_message = f'Error analyzing files: {str(e)}'
            self.status = 'error'
            if self.error_callback:
                self.error_callback('Error', self.error_message)

    def load_excel_data(self):
        """載入 Excel 資料並執行完整分析流程"""
        try:
            self.update_progress(0, 'Starting analysis...')
            self.file_data = FileData()
            
            first_file = self.xlsx_files[0]
            self.file_data.standard_file_name = os.path.basename(first_file)
            self.update_progress(5, "Loading basic info...")
            print(f"Loading basic info from: {os.path.basename(first_file)}")
            
            excel_file = pd.ExcelFile(first_file)
            first_sheet_name = excel_file.sheet_names[0]
            df = pd.read_excel(excel_file, sheet_name=first_sheet_name, header=None)
            
            # 讀取基本資訊
            search_items = {
                'Cust Prod ID': 'cust_prod_id',
                'Cust Prod Group': 'cust_prod_group',
                'Prod Group': 'prod_group',
                'Notch/Flat': 'wafer_notch'
            }
            
            missing_fields = []
            for search_text, attr_name in search_items.items():
                try:
                    row_idx = self.find_cell_in_column_a(df, search_text, self.file_data.standard_file_name)
                    value = df.iloc[row_idx, 1]
                    
                    if pd.isna(value):
                        raise ValueError(f"Value in column B for '{search_text}' is empty")
                        
                    setattr(self.file_data, attr_name, str(value))
                    print(f"  Found '{search_text}' at row {row_idx + 1}, value: {getattr(self.file_data, attr_name)}")
                    
                except ValueError as e:
                    missing_fields.append(search_text)
                    print(f"  Error: {str(e)}")
                except Exception as e:
                    self.error_message = f"Error reading '{search_text}': {str(e)}"
                    self.status = 'error'
                    if self.error_callback:
                        self.error_callback('Data Reading Error', self.error_message)
                    return
            
            if missing_fields:
                error_msg = f"Missing required fields in {self.file_data.standard_file_name}:\n"
                error_msg += "\n".join([f"• {field}" for field in missing_fields])
                error_msg += "\n\nPlease check the Excel file format."
                
                self.error_message = error_msg
                self.status = 'error'
                if self.error_callback:
                    self.error_callback('Missing Required Fields', error_msg)
                return
            
            self.get_wafer_map_range(df)
            self.analyze_wafer_map(first_file)
            
            # 快取座標範圍
            if self.file_data.wafer_map_die_point:
                self.file_data.min_x = min(x for x, y in self.file_data.wafer_map_die_point)
                self.file_data.max_x = max(x for x, y in self.file_data.wafer_map_die_point)
                self.file_data.min_y = min(y for x, y in self.file_data.wafer_map_die_point)
                self.file_data.max_y = max(y for x, y in self.file_data.wafer_map_die_point)
                print(f"  Coordinate range cached: X({self.file_data.min_x}-{self.file_data.max_x}), Y({self.file_data.min_y}-{self.file_data.max_y})")
            
            print(f"Basic Info Loaded:")
            print(f"  Standard file: {self.file_data.standard_file_name}")
            print(f"  Customer Product ID: {self.file_data.cust_prod_id}")
            print(f"  Customer Product Group: {self.file_data.cust_prod_group}")
            print(f"  Product Group: {self.file_data.prod_group}")
            print(f"  Wafer Notch: {self.file_data.wafer_notch}")
            print(f"  Wafer Map X Range: {self.file_data.wafer_map_x_range}")
            print(f"  Wafer Map Y Range: {self.file_data.wafer_map_y_range}")
            print(f"  Total die count: {self.file_data.wafer_map_die_count}")
            print(f"  Sample die coordinates: {self.file_data.wafer_map_die_point[:5]}")
            
            self.update_progress(20, 'Basic info loaded successfully')
            
            if self.should_stop:
                return
            
            print("\nAnalyzing anomalies in all files...")
            self.update_progress(20, "Analyzing anomalies in all files...")
            self.analyze_all_files_anomalies()
            
            if self.should_stop:
                return
            
            self.update_progress(100, 'Analysis complete!')
            self.status = 'completed'
            if self.complete_callback:
                self.complete_callback(self.file_data)
            
        except Exception as e:
            self.error_message = f'Error loading Excel data: {str(e)}'
            self.status = 'error'
            if self.error_callback:
                self.error_callback('Error', self.error_message)
            print(f"Error details: {str(e)}")

    def run(self):
        """執行緒主要執行函數"""
        try:
            self.status = 'running'
            self.load_excel_data()
        except Exception as e:
            self.error_message = f'Analysis failed: {str(e)}'
            self.status = 'error'
            if self.error_callback:
                self.error_callback('Error', self.error_message)
            print(f"Thread error: {str(e)}")


# ==================== 圖表生成函數 ====================

def load_anomaly_weights(process_type='CP'):
    """載入異常代碼權重配置"""
    try:
        base_path = r'M:\BI_Database\Apps\Database\Apps_Database\Q_All\Wafer_Map_Stack_Analysis'
        config_path = os.path.join(base_path, f'{process_type}_Anomaly_weights.json')
        
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            code_to_priority = {}
            for priority, codes in data.get('priority_groups', {}).items():
                priority_int = int(priority)
                for code in codes:
                    code_to_priority[str(code)] = priority_int
            
            return code_to_priority
        else:
            print(f"Warning: {process_type}_Anomaly_weights.json not found at {config_path}")
            return {}
            
    except Exception as e:
        print(f"Error loading anomaly weights: {str(e)}")
        return {}

def prepare_anomaly_codes_data(file_data, weights):
    """準備異常代碼資訊供前端使用"""
    if not file_data or not file_data.die_anomaly_data:
        return []
    
    # 統計每個異常代碼的出現次數和影響的座標數
    anomaly_stats = {}
    for coord, values in file_data.die_anomaly_data.items():
        for code, count in values.items():
            if code not in anomaly_stats:
                anomaly_stats[code] = {
                    'total_count': 0,
                    'coord_count': 0
                }
            anomaly_stats[code]['total_count'] += count
            anomaly_stats[code]['coord_count'] += 1
    
    # 組裝完整資訊
    codes_info = []
    total_die_count = len(file_data.wafer_map_die_point)
    
    for code, stats in anomaly_stats.items():
        name = file_data.anomaly_code_names.get(code, 'Unknown')
        color = file_data.anomaly_code_colors.get(code, 'FF0000')
        if not color.startswith('#'):
            color = '#' + color
        
        percentage = (stats['total_count'] / total_die_count * 100) if total_die_count > 0 else 0
        
        codes_info.append({
            'code': code,
            'name': name,
            'color': color,
            'count': stats['total_count'],
            'coords': stats['coord_count'],
            'percentage': round(percentage, 1)
        })
    
    # 按權重排序
    def get_sort_key(item):
        code = item['code']
        if code in weights:
            return (0, weights[code])
        else:
            return (1, item['code'])
    
    codes_info.sort(key=get_sort_key)
    
    return codes_info


def generate_anomaly_table(file_data, weights):
    """生成異常統計表格 HTML"""
    if not file_data or not file_data.die_anomaly_data:
        return "<p>No anomaly data available</p>"
    
    anomaly_stats = {}
    for coord, values in file_data.die_anomaly_data.items():
        for code, count in values.items():
            if code not in anomaly_stats:
                anomaly_stats[code] = {
                    'total_count': 0,
                    'coordinates': []
                }
            anomaly_stats[code]['total_count'] += count
            anomaly_stats[code]['coordinates'].append(coord)
    
    def get_sort_key(code):
        if code in weights:
            return (0, weights[code])
        else:
            return (1, code)
    
    sorted_codes = sorted(anomaly_stats.keys(), key=get_sort_key)
    
    table_html = """
    <table>
        <thead>
            <tr>
                <th>Priority</th>
                <th>Anomaly Code</th>
                <th>Name</th>
                <th>Total Count</th>
                <th>Affected Coordinates</th>
                <th>Color</th>
            </tr>
        </thead>
        <tbody>
    """
    
    priority = 1
    for code in sorted_codes:
        name = file_data.anomaly_code_names.get(code, 'Unknown')
        color = file_data.anomaly_code_colors.get(code, 'N/A')
        total_count = anomaly_stats[code]['total_count']
        coord_count = len(anomaly_stats[code]['coordinates'])
        
        color_display = f'<div style="display: inline-block; width: 20px; height: 20px; background-color: #{color}; border: 1px solid #000;"></div>' if color != 'N/A' else 'N/A'
        
        if code in weights:
            priority_display = f"#{priority}"
            priority += 1
        else:
            priority_display = "-"
        
        table_html += f"""
            <tr>
                <td style="text-align: center;">{priority_display}</td>
                <td>{code}</td>
                <td>{name}</td>
                <td>{total_count}</td>
                <td>{coord_count} coordinates</td>
                <td>{color_display}</td>
            </tr>
        """
    
    table_html += """
        </tbody>
    </table>
    """
    
    return table_html


def generate_anomaly_distribution_chart(file_data, anomaly_weights, selected_codes=None, 
                                       x_mask_size=None, y_mask_size=None):
    """生成異常分布圖表（含邊緣直方圖）"""
    try:
        if not file_data or not file_data.wafer_map_die_point:
            return None
        
        # 資料篩選
        filtered_die_anomaly_data = {}
        if selected_codes is not None:
            selected_set = set(str(code) for code in selected_codes)
            
            for coord, anomalies in file_data.die_anomaly_data.items():
                filtered_anomalies = {
                    code: count 
                    for code, count in anomalies.items() 
                    if code in selected_set
                }
                if filtered_anomalies:
                    filtered_die_anomaly_data[coord] = filtered_anomalies
        else:
            filtered_die_anomaly_data = file_data.die_anomaly_data
        
        max_x = file_data.wafer_map_x_range
        max_y = file_data.wafer_map_y_range
        
        hover_text_by_coord = {}
        anomaly_counts = {}
        anomaly_coordinates = {}
        coord_to_main_code = {}
        
        x_anomaly_totals = [0] * (max_x + 1)
        y_anomaly_totals = [0] * (max_y + 1)
        
        all_codes = set()
        for coord_data in filtered_die_anomaly_data.values():
            all_codes.update(coord_data.keys())
        
        # 計算每個座標的主要 code
        for (x, y) in file_data.wafer_map_die_point:
            if (x, y) in filtered_die_anomaly_data:
                anomalies = filtered_die_anomaly_data[(x, y)]
                total_anomaly_count = sum(anomalies.values())
                
                x_anomaly_totals[x] += total_anomaly_count
                y_anomaly_totals[y] += total_anomaly_count
                
                max_count = max(anomalies.values())
                candidates = [code for code, count in anomalies.items() if count == max_count]
                
                if len(candidates) > 1:
                    selected_code = min(candidates, key=lambda c: anomaly_weights.get(c, float('inf')))
                else:
                    selected_code = candidates[0]
                
                coord_to_main_code[(x, y)] = selected_code
                
                for code, count in anomalies.items():
                    if code not in anomaly_counts:
                        anomaly_counts[code] = 0
                        anomaly_coordinates[code] = []
                    anomaly_coordinates[code].append((x, y))
                
                name = file_data.anomaly_code_names.get(selected_code, 'Unknown')
                hover_text = f'X: {x}<br>Y: {y}<br>Main Code: {selected_code}<br>Name: {name}<br>Total Anomalies: {total_anomaly_count}'
                hover_text += '<br>────────────────'
                
                sorted_anomalies = sorted(anomalies.items(), key=lambda x: x[1], reverse=True)
                for code, count in sorted_anomalies:
                    anomaly_name = file_data.anomaly_code_names.get(code, 'Unknown')
                    hover_text += f'<br>Code {code} ({anomaly_name}): {count}'
                
                hover_text_by_coord[(x, y)] = hover_text
            else:
                hover_text_by_coord[(x, y)] = f'X: {x}<br>Y: {y}<br>Status: No Anomaly'
        
        # 計算總出現次數
        actual_anomaly_counts = {}
        for coord, anomalies in filtered_die_anomaly_data.items():
            for code, count in anomalies.items():
                if code not in actual_anomaly_counts:
                    actual_anomaly_counts[code] = 0
                actual_anomaly_counts[code] += count
        
        sorted_anomalies = sorted(actual_anomaly_counts.items(), key=lambda x: x[1], reverse=True)
        
        total_die_count = len(file_data.wafer_map_die_point)
        total_anomaly_count = len(filtered_die_anomaly_data)
        anomaly_percentage = (total_anomaly_count / total_die_count * 100) if total_die_count > 0 else 0
        
        worst_anomaly = None
        worst_anomaly_count = 0
        if sorted_anomalies:
            worst_anomaly = sorted_anomalies[0][0]
            worst_anomaly_count = sorted_anomalies[0][1]
        
        fig = go.Figure()
        
        top_margin = 0.12
        right_margin = 0.12
        
        main_domain_x = [0, 1 - right_margin]
        main_domain_y = [0, 1 - top_margin]
        top_domain_y = [1 - top_margin + 0.01, 1]
        right_domain_x = [1 - right_margin + 0.01, 1]
        right_domain_y = main_domain_y
        
        # Base layer
        base_matrix = [[None] * (max_x + 1) for _ in range(max_y + 1)]
        base_hover_matrix = [[None] * (max_x + 1) for _ in range(max_y + 1)]
        
        for (x, y) in file_data.wafer_map_die_point:
            base_matrix[y][x] = 0
            base_hover_matrix[y][x] = hover_text_by_coord.get((x, y), '')
        
        fig.add_trace(go.Heatmap(
            z=base_matrix,
            showscale=False,
            colorscale=[[0, '#cee0e6'], [1, '#cee0e6']],
            hoverongaps=False,
            hovertext=base_hover_matrix,
            hovertemplate='%{hovertext}<extra></extra>',
            xgap=0.5,
            ygap=0.5,
            name='Wafer Base',
            showlegend=False,
            visible=True,
            legendgroup='base',
            xaxis='x',
            yaxis='y'
        ))
        
        # Anomaly code layers
        for i, (code, count) in enumerate(sorted_anomalies):
            code_matrix = [[None] * (max_x + 1) for _ in range(max_y + 1)]
            code_hover_matrix = [[None] * (max_x + 1) for _ in range(max_y + 1)]
            
            color = file_data.anomaly_code_colors.get(code, 'FF0000')
            if not color.startswith('#'):
                color = '#' + color
            
            for (x, y), main_code in coord_to_main_code.items():
                if main_code == code:
                    code_matrix[y][x] = 1
                    code_hover_matrix[y][x] = hover_text_by_coord.get((x, y), '')
            
            name = file_data.anomaly_code_names.get(code, 'Unknown')
            code_percentage = (count / total_die_count * 100) if total_die_count > 0 else 0
            
            fig.add_trace(go.Heatmap(
                z=code_matrix,
                showscale=False,
                colorscale=[[0, color], [1, color]],
                hoverongaps=False,
                hovertext=code_hover_matrix,
                hovertemplate='%{hovertext}<extra></extra>',
                xgap=0.5,
                ygap=0.5,
                name=f'Code {code} ({name}) - {count} ea ({code_percentage:.1f}%)',
                visible=True,
                showlegend=False,
                legendgroup=f'anomaly_{code}',
                xaxis='x',
                yaxis='y'
            ))
            
            # Number annotations
            text_x = []
            text_y = []
            text_values = []
            text_hover = []
            
            for (x, y), main_code in coord_to_main_code.items():
                if main_code == code:
                    total_count = sum(filtered_die_anomaly_data[(x, y)].values())
                    text_x.append(x)
                    text_y.append(y)
                    text_values.append(str(total_count))
                    text_hover.append(hover_text_by_coord.get((x, y), ''))
            
            if text_x:
                fig.add_trace(go.Scatter(
                    x=text_x,
                    y=text_y,
                    text=text_values,
                    mode='text',
                    textfont=dict(
                        size=8,
                        color='black',
                        family='Microsoft JhengHei'
                    ),
                    hovertext=text_hover,
                    hovertemplate='%{hovertext}<extra></extra>',
                    showlegend=False,
                    visible=True,
                    legendgroup=f'anomaly_{code}',
                    xaxis='x',
                    yaxis='y'
                ))
        
        # Top X-axis histogram
        fig.add_trace(
            go.Bar(
                x=list(range(max_x + 1)),
                y=x_anomaly_totals,
                marker_color='#afced8',
                hovertemplate='X: %{x}<br>Total: %{y}<extra></extra>',
                showlegend=False,
                xaxis='x2',
                yaxis='y2'
            )
        )
        
        # Right Y-axis histogram
        fig.add_trace(
            go.Bar(
                x=y_anomaly_totals,
                y=list(range(max_y + 1)),
                orientation='h',
                marker_color='#afced8',
                hovertemplate='Y: %{y}<br>Total: %{x}<extra></extra>',
                showlegend=False,
                xaxis='x3',
                yaxis='y3'
            )
        )
        
        # Title
        title_text = (f'<b>Anomaly Distribution Map with Marginal Histograms</b><br>' +
                    f'<span style="font-size: 20px;">' +
                    f'<b>Total anomaly coordinates: {total_anomaly_count} ea</b> | ')
        
        if worst_anomaly:
            name = file_data.anomaly_code_names.get(worst_anomaly, 'Unknown')
            title_text += f'<b>Code {worst_anomaly} ({name}) has the most occurrences: {worst_anomaly_count} ea</b> | '
        
        title_text += f'<b>Percentage {anomaly_percentage:.1f}%</b></span>'
        
        x_interval = max(1, round((max_x + 1) / 100))
        y_interval = max(1, round((max_y + 1) / 150))
        
        # 遮罩網格線繪製
        shapes = []
        if x_mask_size and y_mask_size:
            die_set = set(file_data.wafer_map_die_point)
            
            for x_start in range(0, max_x + 1, x_mask_size):
                for y_start in range(0, max_y + 1, y_mask_size):
                    x_end = min(x_start + x_mask_size - 1, max_x)
                    y_end = min(y_start + y_mask_size - 1, max_y)
                    
                    # 檢查區域是否包含 die
                    has_die = any((x, y) in die_set 
                                 for x in range(x_start, x_end + 1) 
                                 for y in range(y_start, y_end + 1))
                    
                    if has_die:
                        shapes.append(dict(
                            type="rect",
                            x0=x_start - 0.5, y0=y_start - 0.5,
                            x1=x_end + 0.5, y1=y_end + 0.5,
                            line=dict(color="rgba(128, 128, 128, 0.4)", width=1),
                            fillcolor="rgba(0, 0, 0, 0)",
                            xref="x", yref="y"
                        ))
        
        fig.update_layout(
            title={
                'text': title_text,
                'y': 0.98,
                'x': 0.5,
                'xanchor': 'center',
                'yanchor': 'top',
                'font': dict(
                    size=28,
                    family="Microsoft JhengHei"
                )
            },
            height=1400,
            autosize=True,
            plot_bgcolor='white',
            paper_bgcolor='white',
            margin=dict(
                t=180,
                b=50,
                l=100,
                r=250,
                autoexpand=True
            ),
            showlegend=False,
            shapes=shapes,
            xaxis=dict(
                title='<b>X coordinates</b>',
                showgrid=True,
                gridcolor='#E5E5E5',
                tickmode='linear',
                tick0=0,
                dtick=x_interval,
                showticklabels=True,
                gridwidth=1,
                range=[-0.5, max_x + 0.5],
                zeroline=False,
                tickfont=dict(size=10, family='Microsoft JhengHei'),
                domain=main_domain_x,
                constrain='domain',
                anchor='y'
            ),
            yaxis=dict(
                title='<b>Y coordinates</b>',
                showgrid=True,
                gridcolor='#E5E5E5',
                tickmode='linear',
                tick0=0,
                dtick=y_interval,
                autorange=False,
                range=[max_y + 0.5, -0.5],
                showticklabels=True,
                gridwidth=1,
                zeroline=False,
                tickfont=dict(size=10, family='Microsoft JhengHei'),
                domain=main_domain_y,
                constrain='domain',
                anchor='x'
            ),
            xaxis2=dict(
                showticklabels=False,
                showgrid=False,
                range=[-0.5, max_x + 0.5],
                domain=main_domain_x,
                anchor='y2',
                matches='x'
            ),
            yaxis2=dict(
                showgrid=False,
                showticklabels=True,
                range=[0, max(x_anomaly_totals) * 1.1] if x_anomaly_totals else [0, 1],
                domain=top_domain_y,
                anchor='x2',
                tickfont=dict(size=10, family='Microsoft JhengHei')
            ),
            xaxis3=dict(
                showgrid=False,
                showticklabels=True,
                range=[0, max(y_anomaly_totals) * 1.1] if y_anomaly_totals else [0, 1],
                domain=right_domain_x,
                anchor='y3',
                tickfont=dict(size=10, family='Microsoft JhengHei')
            ),
            yaxis3=dict(
                showticklabels=False,
                showgrid=False,
                range=[max_y + 0.5, -0.5],
                domain=right_domain_y,
                anchor='x3',
                matches='y'
            )
        )
        
        return fig
        
    except Exception as e:
        print(f"Error generating anomaly distribution chart: {str(e)}")
        return None


def generate_anomaly_heatmap_chart(file_data, selected_codes=None, 
                                   x_mask_size=None, y_mask_size=None):
    """生成異常熱力圖"""
    try:
        if not file_data or not file_data.die_anomaly_data:
            return None
        
        # 資料篩選
        filtered_die_anomaly_data = {}
        if selected_codes is not None:
            selected_set = set(str(code) for code in selected_codes)
            
            for coord, anomalies in file_data.die_anomaly_data.items():
                filtered_anomalies = {
                    code: count 
                    for code, count in anomalies.items() 
                    if code in selected_set
                }
                if filtered_anomalies:
                    filtered_die_anomaly_data[coord] = filtered_anomalies
        else:
            filtered_die_anomaly_data = file_data.die_anomaly_data
        
        max_x = file_data.wafer_map_x_range
        max_y = file_data.wafer_map_y_range
        
        hover_text_by_coord = {}
        anomaly_counts = {}
        anomaly_coordinates = {}
        
        # 計算最大總數
        coord_totals = {}
        max_total_count = 0
        for (x, y), anomalies in filtered_die_anomaly_data.items():
            total = sum(anomalies.values())
            coord_totals[(x, y)] = total
            if total > max_total_count:
                max_total_count = total
        
        # 統計資料
        for (x, y) in file_data.wafer_map_die_point:
            if (x, y) in filtered_die_anomaly_data:
                anomalies = filtered_die_anomaly_data[(x, y)]
                total_anomaly_count = sum(anomalies.values())
                
                for code, count in anomalies.items():
                    if code not in anomaly_counts:
                        anomaly_counts[code] = 0
                        anomaly_coordinates[code] = {}
                    anomaly_counts[code] += 1
                    anomaly_coordinates[code][(x, y)] = count
                
                hover_text = f'X: {x}<br>Y: {y}<br>Total Anomalies: {total_anomaly_count}<br>────────────────'
                sorted_anomalies = sorted(anomalies.items(), key=lambda x: x[1], reverse=True)
                for code, count in sorted_anomalies:
                    name = file_data.anomaly_code_names.get(code, 'Unknown')
                    hover_text += f'<br>Code {code} ({name}): {count}'
                
                hover_text_by_coord[(x, y)] = hover_text
            else:
                hover_text_by_coord[(x, y)] = f'X: {x}<br>Y: {y}<br>Status: No Anomaly'
        
        sorted_anomalies = sorted(anomaly_counts.items(), key=lambda x: x[1], reverse=True)
        
        total_die_count = len(file_data.wafer_map_die_point)
        total_anomaly_coords = len(filtered_die_anomaly_data)
        
        # Top 3 coordinates
        sorted_coords = sorted(coord_totals.items(), key=lambda x: x[1], reverse=True)[:3]
        
        fig = go.Figure()
        
        # Base layer
        base_matrix = [[None] * (max_x + 1) for _ in range(max_y + 1)]
        base_hover_matrix = [[None] * (max_x + 1) for _ in range(max_y + 1)]
        
        for (x, y) in file_data.wafer_map_die_point:
            base_matrix[y][x] = 0
            base_hover_matrix[y][x] = hover_text_by_coord.get((x, y), '')
        
        fig.add_trace(go.Heatmap(
            z=base_matrix,
            colorscale='YlOrRd',
            hoverongaps=False,
            hovertext=base_hover_matrix,
            hovertemplate='%{hovertext}<extra></extra>',
            xgap=0.5,
            ygap=0.5,
            showscale=False,
            name='Wafer Base',
            showlegend=False,
            visible=True,
            legendgroup='base',
            zmin=0,
            zmax=100
        ))
        
        # Anomaly code heatmap layers
        for i, (code, count) in enumerate(sorted_anomalies):
            code_matrix = [[None] * (max_x + 1) for _ in range(max_y + 1)]
            code_hover_matrix = [[None] * (max_x + 1) for _ in range(max_y + 1)]
            
            total_occurrences = 0
            
            for (x, y), anomalies in filtered_die_anomaly_data.items():
                if code in anomalies:
                    total_count = sum(anomalies.values())
                    mapped_value = 7 + (total_count / max_total_count * 93) if max_total_count > 0 else 7
                    code_matrix[y][x] = mapped_value
                    code_hover_matrix[y][x] = hover_text_by_coord.get((x, y), '')
                    total_occurrences += anomalies[code]
            
            name = file_data.anomaly_code_names.get(code, 'Unknown')
            
            fig.add_trace(go.Heatmap(
                z=code_matrix,
                colorscale='YlOrRd',
                hoverongaps=False,
                hovertext=code_hover_matrix,
                hovertemplate='%{hovertext}<extra></extra>',
                xgap=0.5,
                ygap=0.5,
                showscale=False,
                name=f'Code {code} ({name}) - {count} coords, {total_occurrences} total',
                visible=True,
                showlegend=False,
                legendgroup=f'anomaly_{code}',
                zmin=0,
                zmax=100
            ))
        
        # Title
        title_text = '<b>Anomaly Heatmap - Occurrences by Coordinate</b><br>'
        title_text += '<span style="font-size: 20px;">'
        title_text += f'<b>Total {total_anomaly_coords} coordinates with anomalies</b> | '
        
        if sorted_coords:
            title_text += '<b>Overall Top 3: '
            top_3_parts = []
            for i, ((x, y), count) in enumerate(sorted_coords):
                top_3_parts.append(f'#{i+1} ({x},{y}): {count}')
            title_text += ' | '.join(top_3_parts)
            title_text += '</b>'
        
        title_text += '</span>'
        
        x_interval = max(1, round((max_x + 1) / 100))
        y_interval = max(1, round((max_y + 1) / 150))
        
        # 遮罩網格線繪製
        shapes = []
        if x_mask_size and y_mask_size:
            die_set = set(file_data.wafer_map_die_point)
            
            for x_start in range(0, max_x + 1, x_mask_size):
                for y_start in range(0, max_y + 1, y_mask_size):
                    x_end = min(x_start + x_mask_size - 1, max_x)
                    y_end = min(y_start + y_mask_size - 1, max_y)
                    
                    # 檢查區域是否包含 die
                    has_die = any((x, y) in die_set 
                                 for x in range(x_start, x_end + 1) 
                                 for y in range(y_start, y_end + 1))
                    
                    if has_die:
                        shapes.append(dict(
                            type="rect",
                            x0=x_start - 0.5, y0=y_start - 0.5,
                            x1=x_end + 0.5, y1=y_end + 0.5,
                            line=dict(color="rgba(128, 128, 128, 0.4)", width=1),
                            fillcolor="rgba(0, 0, 0, 0)",
                            xref="x", yref="y"
                        ))
        
        fig.update_layout(
            title={
                'text': title_text,
                'y': 0.98,
                'x': 0.5,
                'xanchor': 'center',
                'yanchor': 'top',
                'font': dict(
                    size=28,
                    family="Microsoft JhengHei"
                )
            },
            height=1400,
            autosize=True,
            plot_bgcolor='white',
            margin=dict(
                t=180,
                b=50,
                l=100,
                r=250,
                autoexpand=True
            ),
            shapes=shapes,
            xaxis=dict(
                title='<b>X coordinates</b>',
                showgrid=True,
                gridcolor='#E5E5E5',
                tickmode='linear',
                tick0=0,
                dtick=x_interval,
                showticklabels=True,
                gridwidth=1,
                constrain='domain',
                range=[-0.5, max_x + 0.5],
                zeroline=False,
                tickfont=dict(size=10, family='Microsoft JhengHei')
            ),
            yaxis=dict(
                title='<b>Y coordinates</b>',
                showgrid=True,
                gridcolor='#E5E5E5',
                tickmode='linear',
                tick0=0,
                dtick=y_interval,
                autorange='reversed',
                scaleanchor='x',
                scaleratio=1,
                showticklabels=True,
                gridwidth=1,
                constrain='domain',
                range=[-0.5, max_y + 0.5],
                zeroline=False,
                tickfont=dict(size=10, family='Microsoft JhengHei')
            ),
            showlegend=False
        )
        
        fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='#E5E5E5')
        fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='#E5E5E5')
        
        return fig
        
    except Exception as e:
        print(f"Error generating anomaly heatmap chart: {str(e)}")
        return None


def generate_anomaly_ranking_chart(file_data, selected_codes=None):
    """生成異常座標排名圖表（top 50）"""
    try:
        if not file_data or not file_data.die_anomaly_data:
            return None
        
        # 資料篩選
        filtered_die_anomaly_data = {}
        if selected_codes is not None:
            selected_set = set(str(code) for code in selected_codes)
            
            for coord, anomalies in file_data.die_anomaly_data.items():
                filtered_anomalies = {
                    code: count 
                    for code, count in anomalies.items() 
                    if code in selected_set
                }
                if filtered_anomalies:
                    filtered_die_anomaly_data[coord] = filtered_anomalies
        else:
            filtered_die_anomaly_data = file_data.die_anomaly_data
        
        coord_totals = []
        for (x, y), anomalies in filtered_die_anomaly_data.items():
            total_count = sum(anomalies.values())
            coord_totals.append({
                'coord': f'({x},{y})',
                'x': x,
                'y': y,
                'count': total_count
            })
        
        coord_totals.sort(key=lambda x: x['count'], reverse=True)
        top_coords = coord_totals[:50]
        
        if not top_coords:
            return None
        
        coord_labels = [item['coord'] for item in top_coords]
        counts = [item['count'] for item in top_coords]
        
        min_count = min(counts)
        max_count = max(counts)
        
        # 生成顏色漸層
        colors = []
        for count in counts:
            if max_count > min_count:
                normalized = (count - min_count) / (max_count - min_count)
            else:
                normalized = 1.0
            
            if normalized <= 0.25:
                r = 255
                g = int(255 - (255 - 237) * (normalized / 0.25))
                b = int(204 - (204 - 160) * (normalized / 0.25))
            elif normalized <= 0.5:
                r = 255
                g = int(237 - (237 - 178) * ((normalized - 0.25) / 0.25))
                b = int(160 - (160 - 102) * ((normalized - 0.25) / 0.25))
            elif normalized <= 0.75:
                r = 255
                g = int(178 - (178 - 101) * ((normalized - 0.5) / 0.25))
                b = int(102 - (102 - 71) * ((normalized - 0.5) / 0.25))
            else:
                r = int(255 - (255 - 215) * ((normalized - 0.75) / 0.25))
                g = int(101 - (101 - 48) * ((normalized - 0.75) / 0.25))
                b = int(71 - (71 - 39) * ((normalized - 0.75) / 0.25))
            
            colors.append(f'rgb({r}, {g}, {b})')
        
        fig = go.Figure()
        
        fig.add_trace(go.Bar(
            y=coord_labels,
            x=counts,
            orientation='h',
            marker=dict(
                color=colors,
                line=dict(
                    color='rgba(0, 0, 0, 0.2)',
                    width=1
                )
            ),
            text=counts,
            textposition='outside',
            textfont=dict(
                size=10,
                family='Microsoft JhengHei'
            ),
            hovertemplate='Coordinate: %{y}<br>Total Anomalies: %{x}<extra></extra>'
        ))
        
        total_shown = len(top_coords)
        total_coords = len(filtered_die_anomaly_data)
        total_anomalies = sum(counts)
        
        title_text = '<b>Top 50 Anomaly Coordinates Ranking</b><br>'
        title_text += '<span style="font-size: 20px;">'
        title_text += f'<b>Showing {total_shown} of {total_coords} coordinates</b> | '
        title_text += f'<b>Top 50 anomaly count: {total_anomalies}</b>'
        title_text += '</span>'
        
        x_scale = 0.30
        x_offset = (1 - x_scale) / 2
        
        fig.update_layout(
            title={
                'text': title_text,
                'y': 0.98,
                'x': 0.5,
                'xanchor': 'center',
                'yanchor': 'top',
                'font': dict(
                    size=28,
                    family="Microsoft JhengHei"
                )
            },
            height=1400,
            autosize=True,
            plot_bgcolor='white',
            margin=dict(
                t=150,
                b=50,
                l=100,
                r=150,
                autoexpand=True
            ),
            xaxis=dict(
                title='<b>Total Anomaly Count</b>',
                showgrid=True,
                gridcolor='#E5E5E5',
                gridwidth=1,
                zeroline=True,
                zerolinecolor='#666666',
                zerolinewidth=2,
                tickfont=dict(size=12, family='Microsoft JhengHei'),
                domain=[x_offset, x_offset + x_scale]
            ),
            yaxis=dict(
                title='<b>Coordinate (X,Y)</b>',
                showgrid=False,
                tickfont=dict(size=10, family='Microsoft JhengHei'),
                autorange='reversed'
            ),
            showlegend=False
        )
        
        fig.update_xaxes(range=[0, max(counts) * 1.15])
        
        return fig
        
    except Exception as e:
        print(f"Error generating anomaly ranking chart: {str(e)}")
        return None


def generate_mask_overlay_heatmap(file_data, anomaly_weights, selected_codes=None,
                                  x_mask_size=None, y_mask_size=None):
    """生成遮罩疊加累積熱力圖
    
    Args:
        file_data: FileData 物件
        anomaly_weights: 異常代碼權重字典
        selected_codes: 選中的異常代碼列表（可選）
        x_mask_size: X 方向遮罩大小
        y_mask_size: Y 方向遮罩大小
    
    Returns:
        Plotly Figure 物件或 None
    """
    try:
        # 檢查遮罩設定
        if not x_mask_size or not y_mask_size:
            return None
            
        if not file_data or not file_data.wafer_map_die_point:
            return None
        
        max_x = file_data.wafer_map_x_range
        max_y = file_data.wafer_map_y_range
        
        # 資料篩選
        filtered_die_anomaly_data = {}
        if selected_codes is not None:
            selected_set = set(str(code) for code in selected_codes)
            
            for coord, anomalies in file_data.die_anomaly_data.items():
                filtered_anomalies = {
                    code: count 
                    for code, count in anomalies.items() 
                    if code in selected_set
                }
                if filtered_anomalies:
                    filtered_die_anomaly_data[coord] = filtered_anomalies
        else:
            filtered_die_anomaly_data = file_data.die_anomaly_data
        
        # 初始化累積矩陣
        mask_matrix = [[0] * x_mask_size for _ in range(y_mask_size)]
        mask_anomaly_details = [[{} for _ in range(x_mask_size)] for _ in range(y_mask_size)]
        
        # 統計每個異常代碼
        code_accumulation = {}
        
        # 找出包含晶片的遮罩區域
        die_set = set(file_data.wafer_map_die_point)
        processed_masks = 0
        
        # 遍歷遮罩區域
        for mask_start_x in range(0, max_x + 1, x_mask_size):
            for mask_start_y in range(0, max_y + 1, y_mask_size):
                # 檢查是否包含晶片
                has_die = False
                for rel_y in range(y_mask_size):
                    for rel_x in range(x_mask_size):
                        abs_x = mask_start_x + rel_x
                        abs_y = mask_start_y + rel_y
                        if abs_x <= max_x and abs_y <= max_y and (abs_x, abs_y) in die_set:
                            has_die = True
                            break
                    if has_die:
                        break
                
                if not has_die:
                    continue
                
                processed_masks += 1
                
                # 處理區域內晶片
                for rel_y in range(y_mask_size):
                    for rel_x in range(x_mask_size):
                        abs_x = mask_start_x + rel_x
                        abs_y = mask_start_y + rel_y
                        
                        if abs_x <= max_x and abs_y <= max_y and (abs_x, abs_y) in die_set:
                            if (abs_x, abs_y) in filtered_die_anomaly_data:
                                anomalies = filtered_die_anomaly_data[(abs_x, abs_y)]
                                total_count = sum(anomalies.values())
                                
                                # 累加
                                mask_matrix[rel_y][rel_x] += total_count
                                
                                # 累加異常代碼
                                for code, count in anomalies.items():
                                    if code not in code_accumulation:
                                        code_accumulation[code] = [[0] * x_mask_size for _ in range(y_mask_size)]
                                    
                                    code_accumulation[code][rel_y][rel_x] += count
                                    
                                    if code not in mask_anomaly_details[rel_y][rel_x]:
                                        mask_anomaly_details[rel_y][rel_x][code] = 0
                                    mask_anomaly_details[rel_y][rel_x][code] += count
        
        # 找最大累積值
        max_accumulation = 0
        for row in mask_matrix:
            max_accumulation = max(max_accumulation, max(row) if row else 0)
        
        if max_accumulation == 0:
            return None
        
        # 建立 hover text
        hover_texts = []
        for rel_y in range(y_mask_size):
            row_texts = []
            for rel_x in range(x_mask_size):
                total = mask_matrix[rel_y][rel_x]
                if total > 0:
                    hover = f'Relative Position: ({rel_x}, {rel_y})<br>Total Cumulative: {total}<br>────────────────'
                    
                    if mask_anomaly_details[rel_y][rel_x]:
                        sorted_codes = sorted(mask_anomaly_details[rel_y][rel_x].items(), 
                                            key=lambda x: x[1], reverse=True)
                        for code, count in sorted_codes:
                            name = file_data.anomaly_code_names.get(code, 'Unknown')
                            hover += f'<br>Code {code} ({name}): {count}'
                else:
                    hover = f'Relative Position: ({rel_x}, {rel_y})<br>No Anomaly'
                
                row_texts.append(hover)
            hover_texts.append(row_texts)
        
        # 建立圖表
        fig = go.Figure()
        
        # 主熱力圖
        fig.add_trace(go.Heatmap(
            z=mask_matrix,
            colorscale='YlOrRd',
            hovertext=hover_texts,
            hovertemplate='%{hovertext}<extra></extra>',
            colorbar=dict(
                title=dict(
                    text='<b>Cumulative<br>Anomaly<br>Count</b>',
                    font=dict(size=14, family='Microsoft JhengHei')
                ),
                thickness=20,
                len=0.7,
                x=1.02
            ),
            xgap=2,
            ygap=2
        ))
        
        # 標註數字
        text_annotations = []
        for rel_y in range(y_mask_size):
            for rel_x in range(x_mask_size):
                value = mask_matrix[rel_y][rel_x]
                if value > 0:
                    text_annotations.append(
                        dict(
                            x=rel_x,
                            y=rel_y,
                            text=str(value),
                            showarrow=False,
                            font=dict(
                                size=10,
                                color='black' if value < max_accumulation * 0.5 else 'white',
                                family='Microsoft JhengHei'
                            )
                        )
                    )
        
        # Title
        title_text = '<b>Mask Overlay Cumulative Heatmap</b><br>'
        title_text += '<span style="font-size: 20px;">'
        title_text += f'<b>Mask Size: {x_mask_size} x {y_mask_size}</b> | '
        title_text += f'<b>Processed Masks: {processed_masks}</b> | '
        title_text += f'<b>Max Cumulative: {max_accumulation}</b>'
        title_text += '</span>'
        
        fig.update_layout(
            title={
                'text': title_text,
                'y': 0.98,
                'x': 0.5,
                'xanchor': 'center',
                'yanchor': 'top',
                'font': dict(
                    size=28,
                    family="Microsoft JhengHei"
                )
            },
            height=800,
            autosize=True,
            plot_bgcolor='white',
            margin=dict(
                t=150,
                b=50,
                l=100,
                r=200,
                autoexpand=True
            ),
            xaxis=dict(
                title='<b>Relative X Position</b>',
                showgrid=False,
                tickmode='linear',
                tick0=0,
                dtick=1,
                showticklabels=True,
                tickfont=dict(size=10, family='Microsoft JhengHei'),
                range=[-0.5, x_mask_size - 0.5]
            ),
            yaxis=dict(
                title='<b>Relative Y Position</b>',
                showgrid=False,
                tickmode='linear',
                tick0=0,
                dtick=1,
                autorange='reversed',
                showticklabels=True,
                tickfont=dict(size=10, family='Microsoft JhengHei'),
                range=[-0.5, y_mask_size - 0.5]
            ),
            annotations=text_annotations
        )
        
        return fig
        
    except Exception as e:
        print(f"Error generating mask overlay heatmap: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def create_html_report(file_data, weights):
    """建立完整的 HTML 分析報告"""
    cust_prod_id = file_data.cust_prod_id if file_data else 'N/A'
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    total_files = file_data.total_files_processed if file_data else 0
    total_sheets = file_data.total_sheets_processed if file_data else 0
    total_anomalies = len(file_data.die_anomaly_data) if file_data else 0
    
    total_anomaly_count = 0
    if file_data and file_data.die_anomaly_data:
        for coord, values in file_data.die_anomaly_data.items():
            for code, count in values.items():
                total_anomaly_count += count
    
    anomaly_table_html = generate_anomaly_table(file_data, weights)
    
    # 準備異常代碼資料供前端使用
    anomaly_codes_data = prepare_anomaly_codes_data(file_data, weights)
    anomaly_codes_json = json.dumps(anomaly_codes_data)

    # 生成 Plotly.js 內嵌程式碼（唯一載入點，離線可用）
    plotly_js_code = plotly.offline.get_plotlyjs()
    
    # 生成初始圖表（顯示所有 codes）
    anomaly_chart_html = ""
    anomaly_fig = generate_anomaly_distribution_chart(file_data, weights)
    if anomaly_fig:
        anomaly_chart_html = anomaly_fig.to_html(
            include_plotlyjs=False, 
            full_html=False, 
            config={
                "responsive": True,
                "displayModeBar": True,
                "toImageButtonOptions": {
                    "format": "png",
                    "filename": "wafer_map_analysis"
                }
            },
            div_id="plotly-div"
        )
    
    anomaly_heatmap_html = ""
    heatmap_fig = generate_anomaly_heatmap_chart(file_data)
    if heatmap_fig:
        anomaly_heatmap_html = heatmap_fig.to_html(
            include_plotlyjs=False,
            full_html=False,
            config={
                "responsive": True,
                "displayModeBar": True,
                "toImageButtonOptions": {
                    "format": "png",
                    "filename": "anomaly_heatmap"
                }
            },
            div_id="heatmap-div"
        )
    
    anomaly_ranking_html = ""
    ranking_fig = generate_anomaly_ranking_chart(file_data)
    if ranking_fig:
        anomaly_ranking_html = ranking_fig.to_html(
            include_plotlyjs=False,
            full_html=False,
            config={
                "responsive": True,
                "displayModeBar": True,
                "toImageButtonOptions": {
                    "format": "png",
                    "filename": "anomaly_ranking"
                }
            },
            div_id="ranking-div"
        )
    
    return f'''
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Wafer Map Stack Analysis</title>
            <link rel="stylesheet" href="/assets/Google_Fonts/css/noto-sans-tc.css">

            <script type="text/javascript">
                {plotly_js_code}
            </script>

            <style>
                body {{
                    font-family: "Noto Sans TC", Arial, sans-serif;
                    margin: 0;
                    padding: 0;
                    background-color: #f8f9fa;
                }}
                .container {{
                    width: 98%;
                    margin: 20px auto;
                }}
                .header {{
                    background-color: #2D2D2D;
                    color: #E0E0E0;
                    padding: 15px;
                    text-align: center;
                    border-radius: 8px 8px 0 0;
                    margin-bottom: -10px;
                }}
                .tabs {{
                    display: flex;
                    background-color: #2D2D2D;
                    padding: 10px 10px 0 10px;
                    border-radius: 8px 8px 0 0;
                }}
                .tab {{
                    padding: 10px 20px;
                    cursor: pointer;
                    background-color: #454545;
                    color: #E0E0E0;
                    margin-right: 5px;
                    border-radius: 8px 8px 0 0;
                    font-weight: bold;
                }}
                .tab.active {{
                    background-color: #f8f9fa;
                    color: #333;
                }}
                .tab-content {{
                    display: none;
                    padding: 20px;
                    background-color: #fff;
                    border-radius: 0 0 8px 8px;
                    box-shadow: 0 2px 5px rgba(0,0,0,0.1);
                }}
                .tab-content.active {{
                    display: block;
                }}
                .info-section {{
                    margin-bottom: 30px;
                    padding: 15px;
                    background-color: white;
                    border-radius: 8px;
                    box-shadow: 0 2px 5px rgba(0,0,0,0.1);
                    width: 95%;
                    max-width: 1100px;
                    margin-left: auto;
                    margin-right: auto;
                }}
                .chart-container {{
                    margin-bottom: 30px;
                    padding: 15px;
                    background-color: white;
                    border-radius: 8px;
                    box-shadow: 0 2px 5px rgba(0,0,0,0.1);
                    width: 98%;
                    margin-left: auto;
                    margin-right: auto;
                    overflow-x: auto;
                    overflow-y: hidden;
                }}
                #plotly-div, #heatmap-div, #ranking-div, #mask-overlay-div {{
                    width: 100%;
                    height: 100%;
                }}
                .js-plotly-plot {{
                    width: 100% !important;
                    height: auto !important;
                }}
                .plotly-graph-div {{
                    width: 100% !important;
                    margin: 0 auto !important;
                }}
                .info-title {{
                    font-size: 20px;
                    font-weight: bold;
                    margin-bottom: 15px;
                    color: #333;
                    border-bottom: 2px solid #e0e0e0;
                    padding-bottom: 10px;
                }}
                .info-grid {{
                    display: grid;
                    grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
                    gap: 15px;
                    margin-bottom: 20px;
                }}
                .info-item {{
                    padding: 10px;
                    background-color: #f8f9fa;
                    border-radius: 5px;
                    border-left: 3px solid #4A4A4A;
                }}
                .info-label {{
                    font-weight: bold;
                    color: #666;
                    margin-bottom: 5px;
                }}
                .info-value {{
                    color: #333;
                    font-size: 16px;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 10px;
                }}
                th, td {{
                    border: 1px solid #ddd;
                    padding: 8px;
                    text-align: left;
                }}
                th {{
                    background-color: #2D2D2D;
                    color: white;
                    font-weight: bold;
                }}
                tr:nth-child(even) {{
                    background-color: #f8f9fa;
                }}
                .summary-box {{
                    background-color: #D4EDDA;
                    border-left: 4px solid #28A745;
                    padding: 15px;
                    margin: 20px 0;
                    border-radius: 5px;
                }}
                .note-box {{
                    background-color: #f5f5f5;
                    border-left: 4px solid #4A4A4A;
                    padding: 12px;
                    margin: 15px 0;
                    border-radius: 5px;
                    font-size: 14px;
                    color: #666;
                }}
                
                /* Unified Control Panel Styles */
                .unified-control-panel {{
                    margin: 20px auto 60px auto;
                    width: 98%;
                    background: white;
                    border-radius: 12px;
                    border: 2px solid #e9ecef;
                    overflow: hidden;
                    box-shadow: 0 4px 12px rgba(0,0,0,0.08);
                }}
                .control-panel-header {{
                    background: linear-gradient(135deg, #2D2D2D 0%, #4A4A4A 100%);
                    color: white;
                    padding: 18px 24px;
                    display: flex;
                    justify-content: space-between;
                    align-items: center;
                }}
                .header-left {{
                    display: flex;
                    align-items: center;
                    gap: 12px;
                }}
                .header-icon {{
                    font-size: 24px;
                }}
                .header-title {{
                    font-size: 18px;
                    font-weight: 600;
                    letter-spacing: 0.3px;
                }}
                .header-subtitle {{
                    font-size: 13px;
                    opacity: 0.85;
                    margin-top: 4px;
                }}
                .control-panel-body {{
                    padding: 24px;
                    max-height: 600px;
                    overflow-y: auto;
                    background: #fafafa;
                }}
                
                /* Mask Settings Section */
                .mask-settings-section {{
                    margin-bottom: 30px;
                    padding-bottom: 30px;
                    border-bottom: 2px solid #e9ecef;
                }}
                .mask-title {{
                    font-size: 16px;
                    font-weight: 700;
                    color: #2D2D2D;
                    margin-bottom: 15px;
                    display: flex;
                    align-items: center;
                    gap: 8px;
                }}
                .mask-inputs {{
                    display: flex;
                    gap: 20px;
                    align-items: center;
                    flex-wrap: wrap;
                }}
                .input-group {{
                    display: flex;
                    flex-direction: column;
                    gap: 8px;
                }}
                .input-label {{
                    font-size: 13px;
                    font-weight: 600;
                    color: #666;
                }}
                .mask-input {{
                    width: 120px;
                    padding: 10px 12px;
                    border: 2px solid #e9ecef;
                    border-radius: 8px;
                    font-size: 15px;
                    font-family: "Noto Sans TC", Arial, sans-serif;
                    transition: all 0.3s ease;
                }}
                .mask-input:focus {{
                    outline: none;
                    border-color: #4A4A4A;
                    box-shadow: 0 0 0 3px rgba(74, 74, 74, 0.1);
                }}
                .mask-info {{
                    font-size: 12px;
                    color: #6c757d;
                    margin-top: 10px;
                }}
                
                /* Code Filter Section */
                .code-filter-section {{
                    margin-bottom: 20px;
                }}
                .section-title {{
                    font-size: 16px;
                    color: #2D2D2D;
                    margin-bottom: 15px;
                    font-weight: 700;
                    display: flex;
                    align-items: center;
                    gap: 8px;
                }}
                
                /* Quick Actions */
                .quick-actions {{
                    display: flex;
                    gap: 10px;
                    flex-wrap: wrap;
                }}
                .quick-btn {{
                    padding: 8px 18px;
                    border: none;
                    background: rgba(255,255,255,0.15);
                    border-radius: 8px;
                    cursor: pointer;
                    font-size: 13px;
                    font-weight: 600;
                    color: white;
                    transition: all 0.2s ease;
                    display: flex;
                    align-items: center;
                    gap: 6px;
                }}
                .quick-btn:hover {{
                    background: rgba(255,255,255,0.25);
                    border-color: rgba(255,255,255,0.5);
                    transform: translateY(-1px);
                }}
                .quick-btn.apply-filter {{
                    background: linear-gradient(135deg, #4A4A4A 0%, #2C2C2C 100%);
                    border: none;
                    box-shadow: 0 2px 8px rgba(74, 74, 74, 0.3);
                    position: relative;
                    padding-left: 28px;
                }}
                .quick-btn.apply-filter::before {{
                    content: '';
                    position: absolute;
                    left: 10px;
                    top: 50%;
                    transform: translateY(-50%);
                    width: 8px;
                    height: 8px;
                    border-radius: 50%;
                    background: linear-gradient(135deg, #D0D0D0 0%, #A0A0A0 100%);
                    box-shadow: 0 0 8px rgba(160, 160, 160, 0.6);
                }}
                .quick-btn.apply-filter:hover:not(:disabled) {{
                    transform: translateY(-2px);
                    box-shadow: 0 4px 12px rgba(74, 74, 74, 0.4);
                    background: linear-gradient(135deg, #5A5A5A 0%, #3A3A3A 100%);
                }}
                .quick-btn.apply-filter:disabled {{
                    background: rgba(255,255,255,0.1);
                    color: rgba(255,255,255,0.4);
                    cursor: not-allowed;
                    transform: none;
                    box-shadow: none;
                }}
                .quick-btn.apply-filter:disabled::before {{
                    background: rgba(255,255,255,0.3);
                    box-shadow: none;
                }}
                
                /* Code Grid */
                .code-grid {{
                    display: flex;
                    flex-wrap: wrap;
                    gap: 12px;
                }}
                .code-card {{
                    flex: 0 1 320px;
                    max-width: 360px;
                    background: white;
                    border: 2px solid #e9ecef;
                    border-radius: 10px;
                    padding: 14px 16px;
                    cursor: pointer;
                    transition: all 0.3s ease;
                    display: flex;
                    align-items: center;
                    gap: 12px;
                    position: relative;
                    overflow: hidden;
                }}
                .code-card:hover {{
                    box-shadow: 0 4px 12px rgba(0,0,0,0.1);
                    transform: translateY(-2px);
                }}
                .code-card.selected {{
                    box-shadow: 0 2px 8px rgba(var(--card-color-rgb), 0.15);
                }}
                .code-checkbox {{
                    width: 22px;
                    height: 22px;
                    cursor: pointer;
                    flex-shrink: 0;
                }}
                .code-color-indicator {{
                    width: 32px;
                    height: 32px;
                    border-radius: 8px;
                    border: 2px solid #dee2e6;
                    flex-shrink: 0;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                }}
                .code-info {{
                    flex: 1;
                    min-width: 0;
                    text-align: left;
                }}
                .code-header {{
                    display: flex;
                    justify-content: space-between;
                    align-items: center;
                    margin-bottom: 6px;
                }}
                .code-label {{
                    font-weight: 700;
                    color: #2D2D2D;
                    font-size: 15px;
                }}
                .priority-badge {{
                    background: linear-gradient(135deg, #4A4A4A 0%, #2D2D2D 100%);
                    color: white;
                    padding: 2px 10px;
                    border-radius: 12px;
                    font-size: 11px;
                    font-weight: 700;
                    letter-spacing: 0.5px;
                }}
                .code-name {{
                    color: #6c757d;
                    font-size: 13px;
                    margin-bottom: 6px;
                    white-space: nowrap;
                    overflow: hidden;
                    text-overflow: ellipsis;
                    text-align: left;
                }}
                .code-stats {{
                    display: flex;
                    gap: 12px;
                    font-size: 12px;
                    color: #868e96;
                    flex-wrap: wrap;
                    justify-content: flex-start;
                }}
                .stat-badge {{
                    display: flex;
                    align-items: center;
                    gap: 4px;
                    font-weight: normal;
                }}
                .stat-badge .value {{
                    color: #2D2D2D;
                    font-weight: normal;
                }}
                
                /* Scrollbar Styling */
                .control-panel-body::-webkit-scrollbar {{
                    width: 10px;
                }}
                .control-panel-body::-webkit-scrollbar-track {{
                    background: #f1f1f1;
                    border-radius: 10px;
                }}
                .control-panel-body::-webkit-scrollbar-thumb {{
                    background: #c1c1c1;
                    border-radius: 10px;
                }}
                .control-panel-body::-webkit-scrollbar-thumb:hover {{
                    background: #a8a8a8;
                }}
                
                /* Loading State */
                .loading-overlay {{
                    display: none;
                    position: fixed;
                    top: 0;
                    left: 0;
                    right: 0;
                    bottom: 0;
                    background: rgba(0, 0, 0, 0.6);
                    backdrop-filter: blur(4px);
                    z-index: 9999;
                    justify-content: center;
                    align-items: center;
                }}
                .loading-overlay.active {{
                    display: flex;
                }}
                .loading-content {{
                    background: white;
                    padding: 40px;
                    border-radius: 16px;
                    text-align: center;
                    box-shadow: 0 8px 32px rgba(0,0,0,0.3);
                }}
                .spinner {{
                    border: 4px solid #f3f3f3;
                    border-top: 4px solid #2D2D2D;
                    border-radius: 50%;
                    width: 50px;
                    height: 50px;
                    animation: spin 1s linear infinite;
                    margin: 0 auto 20px;
                }}
                @keyframes spin {{
                    0% {{ transform: rotate(0deg); }}
                    100% {{ transform: rotate(360deg); }}
                }}
                .loading-text {{
                    font-size: 16px;
                    font-weight: 600;
                    color: #2D2D2D;
                    margin-bottom: 8px;
                }}
                .loading-subtext {{
                    font-size: 14px;
                    color: #6c757d;
                }}
                
                /* Confirm Modal Styles */
                .modal {{
                    display: none;
                    position: fixed;
                    top: 0;
                    left: 0;
                    width: 100%;
                    height: 100%;
                    background: rgba(0, 0, 0, 0.45);
                    backdrop-filter: blur(4px);
                    justify-content: center;
                    align-items: center;
                    z-index: 10000;
                }}
                .modal.show {{
                    display: flex;
                    animation: fadeIn 0.3s ease;
                }}
                @keyframes fadeIn {{
                    from {{
                        opacity: 0;
                    }}
                    to {{
                        opacity: 1;
                    }}
                }}
                .modal-content {{
                    background: white;
                    border-radius: 16px;
                    padding: 32px;
                    max-width: 400px;
                    width: 90%;
                    text-align: center;
                    animation: slideUp 0.3s ease;
                    box-shadow: 0 12px 48px rgba(0, 0, 0, 0.15);
                }}
                @keyframes slideUp {{
                    from {{
                        opacity: 0;
                        transform: translateY(50px);
                    }}
                    to {{
                        opacity: 1;
                        transform: translateY(0);
                    }}
                }}
                .modal-icon {{
                    font-size: 48px;
                    margin-bottom: 20px;
                    color: #4A4A4A;
                }}
                .modal-title {{
                    font-size: 20px;
                    font-weight: 700;
                    margin-bottom: 12px;
                    color: #2D2D2D;
                }}
                .modal-message {{
                    font-size: 14px;
                    color: #666666;
                    margin-bottom: 24px;
                    white-space: pre-line;
                    line-height: 1.6;
                }}
                .modal-buttons {{
                    display: flex;
                    gap: 10px;
                    justify-content: center;
                }}
                .modal-btn {{
                    padding: 12px 24px;
                    border: none;
                    border-radius: 10px;
                    font-size: 14px;
                    font-weight: 600;
                    cursor: pointer;
                    transition: all 0.3s ease;
                    background: linear-gradient(135deg, #4A4A4A 0%, #2C2C2C 100%);
                    color: white;
                    letter-spacing: 0.3px;
                    flex: 1;
                }}
                .modal-btn:hover {{
                    transform: translateY(-2px);
                    box-shadow: 0 6px 16px rgba(44, 44, 44, 0.25);
                }}
                .modal-btn-secondary {{
                    background: #E8E8E8;
                    color: #2D2D2D;
                }}
                .modal-btn-secondary:hover {{
                    background: #D8D8D8;
                    box-shadow: 0 6px 16px rgba(0, 0, 0, 0.1);
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>Wafer Map Stack Analysis</h1>
                    <p style="font-size: 14px;">Customer Product ID: {cust_prod_id}</p>
                    <p style="font-size: 14px;">Analysis Time: {timestamp}</p>
                </div>
                
                <div class="tabs">
                    <div class="tab active" onclick="showTab('info')">Info</div>
                    <div class="tab" onclick="showTab('analysis')">Analysis</div>
                </div>
                
                <div id="info" class="tab-content active">
                    <div class="info-section">
                        <div class="info-title">Basic Information</div>
                        <div class="info-grid">
                            <div class="info-item">
                                <div class="info-label">Customer Product ID</div>
                                <div class="info-value">{file_data.cust_prod_id if file_data else 'N/A'}</div>
                            </div>
                            <div class="info-item">
                                <div class="info-label">Customer Product Group</div>
                                <div class="info-value">{file_data.cust_prod_group if file_data else 'N/A'}</div>
                            </div>
                            <div class="info-item">
                                <div class="info-label">Product Group</div>
                                <div class="info-value">{file_data.prod_group if file_data else 'N/A'}</div>
                            </div>
                            <div class="info-item">
                                <div class="info-label">Wafer Notch</div>
                                <div class="info-value">{file_data.wafer_notch if file_data else 'N/A'}</div>
                            </div>
                        </div>
                    </div>
                    
                    <div class="info-section">
                        <div class="info-title">Wafer Map Information</div>
                        <div class="info-grid">
                            <div class="info-item">
                                <div class="info-label">X Range</div>
                                <div class="info-value">0 - {file_data.wafer_map_x_range if file_data else 'N/A'}</div>
                            </div>
                            <div class="info-item">
                                <div class="info-label">Y Range</div>
                                <div class="info-value">0 - {file_data.wafer_map_y_range if file_data else 'N/A'}</div>
                            </div>
                            <div class="info-item">
                                <div class="info-label">Total Die Count</div>
                                <div class="info-value">{file_data.wafer_map_die_count if file_data else 'N/A'}</div>
                            </div>
                        </div>
                    </div>
                    
                    <div class="info-section">
                        <div class="info-title">Analysis Summary</div>
                        <div class="summary-box">
                            <p><strong>Total Files Processed:</strong> {total_files}</p>
                            <p><strong>Total Sheets Processed:</strong> {total_sheets}</p>
                            <p><strong>Total Anomaly Coordinates:</strong> {total_anomalies}</p>
                            <p><strong>Total Anomaly Count:</strong> {total_anomaly_count}</p>
                        </div>
                        
                        <div class="info-title">Anomaly Details</div>
                        <div class="note-box">
                            <strong>Note:</strong>
                            <br>• <strong>Total Count</strong> - The total number of times this anomaly appears across all files (including duplicates at the same coordinate)
                            <br>• <strong>Affected Coordinates</strong> - The number of unique coordinate locations where this anomaly was detected
                        </div>
                        {anomaly_table_html}
                    </div>
                </div>
                
                <div id="analysis" class="tab-content">
                    <!-- Unified Control Panel -->
                    <div class="unified-control-panel">
                        <div class="control-panel-header">
                            <div class="header-left">
                                <span class="header-icon">⚙</span>
                                <div>
                                    <div class="header-title">Analysis Control Panel</div>
                                    <div class="header-subtitle">Configure anomaly code filter and mask overlay settings</div>
                                </div>
                            </div>
                            <div class="quick-actions">
                                <button class="quick-btn select-all" onclick="selectAllCodes()">
                                    ✓ Select All
                                </button>
                                <button class="quick-btn clear-all" onclick="clearAllCodes()">
                                    ✕ Clear All
                                </button>
                                <button class="quick-btn apply-filter" id="applyFilterBtn" onclick="applyAllFilters()">
                                    <span id="btnText">Apply & Regenerate Charts</span>
                                </button>
                            </div>
                        </div>
                        
                        <div class="control-panel-body" id="controlPanelBody">
                            <!-- Mask Settings Section -->
                            <div class="mask-settings-section">
                                <div class="mask-title">
                                    📐 Mask Overlay Settings
                                </div>
                                <div class="mask-inputs">
                                    <div class="input-group">
                                        <label class="input-label">X Mask Size</label>
                                        <input type="number" class="mask-input" id="xMaskInput" 
                                               min="1" max="100" placeholder="e.g. 10">
                                    </div>
                                    <div class="input-group">
                                        <label class="input-label">Y Mask Size</label>
                                        <input type="number" class="mask-input" id="yMaskInput" 
                                               min="1" max="100" placeholder="e.g. 10">
                                    </div>
                                </div>
                                <div class="mask-info">
                                    Enter mask size to enable overlay grid lines and generate cumulative heatmap (Range: 1-100)
                                </div>
                            </div>
                            
                            <!-- Code Filter Section -->
                            <div class="code-filter-section">
                                <div class="section-title">
                                    ⚙️ Anomaly Code Filter
                                </div>
                                <div class="code-grid" id="codeGrid">
                                </div>
                            </div>
                        </div>
                    </div>
                    
                    <div class="chart-container" id="anomaly-distribution">
                        {anomaly_chart_html}
                    </div>
                    
                    <div class="chart-container" id="anomaly-heatmap">
                        {anomaly_heatmap_html}
                    </div>
                    
                    <div class="chart-container" id="anomaly-ranking">
                        {anomaly_ranking_html}
                    </div>
                    
                    <div class="chart-container" id="mask-overlay-container" style="display: none;">
                        <div id="mask-overlay-div"></div>
                    </div>
                </div>
            </div>
            
            <div class="loading-overlay" id="loadingOverlay">
                <div class="loading-content">
                    <div class="spinner"></div>
                    <div class="loading-text">Regenerating Charts...</div>
                    <div class="loading-subtext">Please wait while we update the visualizations</div>
                </div>
            </div>
            
            <!-- Confirm Modal -->
            <div class="modal" id="confirmModal">
                <div class="modal-content">
                    <div class="modal-icon">
                        ❓
                    </div>
                    <div class="modal-title" id="confirmTitle">Confirm</div>
                    <div class="modal-message" id="confirmMessage">Message</div>
                    <div class="modal-buttons">
                        <button class="modal-btn modal-btn-secondary" id="confirmCancelBtn">Cancel</button>
                        <button class="modal-btn" id="confirmOkBtn">Confirm</button>
                    </div>
                </div>
            </div>
            
            <script>
                // 瀏覽器關閉主動通知機制
                let isNormalNavigation = false;
                
                window.addEventListener('beforeunload', () => {{
                    if (!isNormalNavigation) {{
                        fetch('/api/shutdown', {{
                            method: 'POST',
                            keepalive: true
                        }});
                    }}
                }});
                
                // 心跳保活機制
                setInterval(() => {{
                    fetch('/api/heartbeat', {{
                        method: 'POST',
                        headers: {{ 'Content-Type': 'application/json' }}
                    }}).catch(error => {{
                        console.error('Heartbeat failed:', error);
                    }});
                }}, 60000);
                
                const anomalyCodesData = {anomaly_codes_json};
                
                const xMaskInput = document.getElementById('xMaskInput');
                const yMaskInput = document.getElementById('yMaskInput');
                const applyFilterBtn = document.getElementById('applyFilterBtn');
                const btnText = document.getElementById('btnText');
                
                // Promise-based 確認對話框
                async function showConfirm(message, isAlert = false) {{
                    return new Promise((resolve) => {{
                        const confirmModal = document.getElementById('confirmModal');
                        const confirmMessage = document.getElementById('confirmMessage');
                        const confirmOkBtn = document.getElementById('confirmOkBtn');
                        const confirmCancelBtn = document.getElementById('confirmCancelBtn');
                        
                        confirmMessage.textContent = message;
                        
                        if (isAlert) {{
                            confirmCancelBtn.style.display = 'none';
                            confirmOkBtn.textContent = 'OK';
                        }} else {{
                            confirmCancelBtn.style.display = 'block';
                            confirmOkBtn.textContent = 'Confirm';
                        }}
                        
                        confirmModal.classList.add('show');
                        
                        confirmOkBtn.onclick = () => {{
                            confirmModal.classList.remove('show');
                            resolve(true);
                        }};
                        
                        confirmCancelBtn.onclick = () => {{
                            confirmModal.classList.remove('show');
                            resolve(false);
                        }};
                    }});
                }}
                
                function hexToRgb(hex) {{
                    const result = /^#?([a-f\\d]{{2}})([a-f\\d]{{2}})([a-f\\d]{{2}})$/i.exec(hex);
                    return result ? 
                        `${{parseInt(result[1], 16)}}, ${{parseInt(result[2], 16)}}, ${{parseInt(result[3], 16)}}` : 
                        '0, 0, 0';
                }}
                
                function initControlPanel() {{
                    const container = document.getElementById('codeGrid');
                    
                    if (!container) return;
                    
                    container.innerHTML = '';
                    
                    anomalyCodesData.forEach((codeInfo, index) => {{
                        const colorRgb = hexToRgb(codeInfo.color);
                        const card = document.createElement('div');
                        card.className = 'code-card selected';
                        card.style.setProperty('--card-color', codeInfo.color);
                        card.style.setProperty('--card-color-rgb', colorRgb);
                        card.onclick = () => toggleCard(card);
                        
                        const priorityBadge = index < 10 ? `<span class="priority-badge">#${{index + 1}}</span>` : '';
                        
                        card.innerHTML = `
                            <input type="checkbox" class="code-checkbox" checked onclick="event.stopPropagation(); toggleCard(this.parentElement)">
                            <div class="code-color-indicator" style="background-color: ${{codeInfo.color}};"></div>
                            <div class="code-info">
                                <div class="code-header">
                                    <span class="code-label">Code ${{codeInfo.code}}</span>
                                    ${{priorityBadge}}
                                </div>
                                <div class="code-name">${{codeInfo.name}}</div>
                                <div class="code-stats">
                                    <span class="stat-badge">
                                        <span class="value">${{codeInfo.count}}</span> Cumulate frequency
                                    </span>
                                </div>
                            </div>
                        `;
                        
                        container.appendChild(card);
                    }});
                }}
                
                function toggleCard(element) {{
                    const card = element.classList.contains('code-card') ? element : element.closest('.code-card');
                    const checkbox = card.querySelector('.code-checkbox');
                    
                    if (element.tagName !== 'INPUT') {{
                        checkbox.checked = !checkbox.checked;
                    }}
                    
                    if (checkbox.checked) {{
                        card.classList.add('selected');
                    }} else {{
                        card.classList.remove('selected');
                    }}
                }}
                
                function selectAllCodes() {{
                    document.querySelectorAll('.code-card').forEach(card => {{
                        card.classList.add('selected');
                        card.querySelector('.code-checkbox').checked = true;
                    }});
                }}
                
                function clearAllCodes() {{
                    document.querySelectorAll('.code-card').forEach(card => {{
                        card.classList.remove('selected');
                        card.querySelector('.code-checkbox').checked = false;
                    }});
                    xMaskInput.value = '';
                    yMaskInput.value = '';
                    updateButtonState();
                }}
                
                function getSelectedCodes() {{
                    const checkboxes = document.querySelectorAll('.code-card.selected .code-checkbox:checked');
                    return Array.from(checkboxes).map((cb, index) => {{
                        const card = cb.closest('.code-card');
                        const cardIndex = Array.from(document.querySelectorAll('.code-card')).indexOf(card);
                        return anomalyCodesData[cardIndex].code;
                    }});
                }}
                
                // 更新按鈕狀態
                function updateButtonState() {{
                    const xValue = xMaskInput.value.trim();
                    const yValue = yMaskInput.value.trim();
                    const hasMask = xValue && yValue;
                    
                    if (hasMask) {{
                        btnText.textContent = 'Apply All & Regenerate (3 Charts)';
                    }} else {{
                        btnText.textContent = 'Apply & Regenerate Charts';
                    }}
                }}
                
                // 監聽輸入變化
                xMaskInput.addEventListener('input', updateButtonState);
                yMaskInput.addEventListener('input', updateButtonState);
                
                // 統一處理函數
                async function applyAllFilters() {{
                    const selectedCodes = getSelectedCodes();
                    
                    if (selectedCodes.length === 0) {{
                        await showConfirm('Please select at least one anomaly code.', true);
                        return;
                    }}
                    
                    const xMask = xMaskInput.value;
                    const yMask = yMaskInput.value;
                    const xNum = xMask ? parseInt(xMask) : null;
                    const yNum = yMask ? parseInt(yMask) : null;
                    
                    let validMask = null;
                    if (xNum !== null && yNum !== null) {{
                        if (xNum > 0 && yNum > 0 && xNum <= 100 && yNum <= 100) {{
                            validMask = {{ x: xNum, y: yNum }};
                        }} else {{
                            await showConfirm('Mask Size must be between 1 and 100.', true);
                            return;
                        }}
                    }}
                    
                    const loadingOverlay = document.getElementById('loadingOverlay');
                    
                    try {{
                        applyFilterBtn.disabled = true;
                        loadingOverlay.classList.add('active');
                        
                        const loadingText = document.querySelector('.loading-text');
                        if (validMask) {{
                            loadingText.textContent = 'Regenerating All Charts (with Mask)...';
                        }} else {{
                            loadingText.textContent = 'Regenerating Charts...';
                        }}
                        
                        const response = await fetch('/api/regenerate_all_charts', {{
                            method: 'POST',
                            headers: {{
                                'Content-Type': 'application/json'
                            }},
                            body: JSON.stringify({{
                                selected_codes: selectedCodes,
                                x_mask_size: validMask ? validMask.x : null,
                                y_mask_size: validMask ? validMask.y : null
                            }})
                        }});
                        
                        const result = await response.json();
                        
                        if (result.success) {{
                            Plotly.react('plotly-div', result.distribution.data, 
                                        result.distribution.layout, result.distribution.config);
                            
                            Plotly.react('heatmap-div', result.heatmap.data, 
                                        result.heatmap.layout, result.heatmap.config);
                            
                            const maskContainer = document.getElementById('mask-overlay-container');
                            if (result.mask_heatmap) {{
                                maskContainer.style.display = 'block';
                                Plotly.react('mask-overlay-div', result.mask_heatmap.data, 
                                            result.mask_heatmap.layout, result.mask_heatmap.config);
                            }} else {{
                                maskContainer.style.display = 'none';
                            }}
                            
                        }} else {{
                            await showConfirm('Error regenerating charts: ' + result.error, true);
                        }}
                        
                    }} catch (error) {{
                        console.error('Error:', error);
                        await showConfirm('Failed to regenerate charts. Please try again.', true);
                    }} finally {{
                        applyFilterBtn.disabled = false;
                        loadingOverlay.classList.remove('active');
                    }}
                }}
                
                function showTab(tabId) {{
                    var contents = document.getElementsByClassName('tab-content');
                    for (var i = 0; i < contents.length; i++) {{
                        contents[i].classList.remove('active');
                    }}
                    
                    var tabs = document.getElementsByClassName('tab');
                    for (var i = 0; i < tabs.length; i++) {{
                        tabs[i].classList.remove('active');
                    }}
                    
                    document.getElementById(tabId).classList.add('active');
                    document.querySelector('.tab[onclick="showTab(\\''+tabId+'\\')"]').classList.add('active');
                    
                    if (tabId === 'analysis') {{
                        window.dispatchEvent(new Event('resize'));
                    }}
                }}
                
                window.addEventListener('load', function() {{
                    initControlPanel();
                    updateButtonState();
                    window.dispatchEvent(new Event('resize'));
                }});
                
                window.addEventListener('click', (e) => {{
                    if (e.target === document.getElementById('confirmModal')) {{
                        document.getElementById('confirmModal').classList.remove('show');
                    }}
                }});
            </script>
        </body>
        </html>
        '''

# ==================== 工具函數 ====================

def update_last_activity():
    """更新最後活動時間"""
    global last_activity_time
    last_activity_time = time.time()


def check_activity_thread():
    """背景執行緒：檢查活動逾時"""
    global last_activity_time
    while True:
        time.sleep(60)
        if time.time() - last_activity_time > ACTIVITY_TIMEOUT:
            print("Activity timeout reached. Shutting down...")
            os._exit(0)


def is_port_in_use(port):
    """檢查指定埠是否被使用"""
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        return s.connect_ex(('localhost', port)) == 0


def find_available_port(start_port=8000, end_port=8999):
    """在指定範圍內尋找可用埠"""
    for port in range(start_port, end_port + 1):
        if not is_port_in_use(port):
            return port
    raise RuntimeError(f"No available ports in range {start_port}-{end_port}")


def save_log():
    """儲存使用記錄到 SQL Server"""
    try:
        current_datetime = datetime.now()
        conn_str = f'DRIVER={{SQL Server}};SERVER={SQL_SERVER_INFO["server"]};DATABASE={SQL_SERVER_INFO["database"]};UID={SQL_SERVER_INFO["username"]};PWD={SQL_SERVER_INFO["password"]};App=Wafer Map Stack Analysis'
        
        with pyodbc.connect(conn_str) as conn:
            cursor = conn.cursor()
            insert_query = f"""
            INSERT INTO {SQL_SERVER_INFO["apps_log_table"]} (Activation_Time, User_Id, Status, Apps_Name)
            VALUES (?, ?, ?, ?)
            """
            cursor.execute(insert_query, (current_datetime, username, "Open", "Wafer Map Stack Analysis"))
            conn.commit()
            
    except Exception as e:
        print(f"Error writing to log database: {str(e)}")


def check_version():
    """檢查版本更新"""
    try:
        app_folder = os.path.normpath(os.path.join("M:", "BI_Database", "Apps", "Database", "Apps_Installation_package", "Q_All"))
        exe_files = [os.path.join(app_folder, f) for f in os.listdir(app_folder) 
                    if f.startswith("Wafer Map Stack Analysis_V") and f.endswith(".exe")]

        if not exe_files:
            return {
                'status': 'error',
                'type': 'permission',
                'message': 'Failed to obtain launch permission'
            }

        def parse_version(version_str):
            match = re.search(r'_V(\d+)\.(\d+)', version_str)
            if match:
                return tuple(map(int, match.groups()))
            return (0, 0)

        latest_version = max(parse_version(os.path.basename(f)) for f in exe_files)
        latest_exe = max((f for f in exe_files 
                        if parse_version(os.path.basename(f)) == latest_version), 
                    key=os.path.getmtime)

        current_version_match = re.search(r'_V(\d+)\.(\d+)', os.path.basename(sys.executable))
        if current_version_match:
            current_version = tuple(map(int, current_version_match.groups()))
        else:
            current_version = (6, 0)

        if current_version[0] != latest_version[0]:
            return {
                'status': 'update',
                'message': f'Will update to new version V{latest_version[0]}.{latest_version[1]}',
                'latest_exe': latest_exe
            }

        if current_version[0] != 6:
            return {
                'status': 'error',
                'type': 'version',
                'message': 'Software version error'
            }

        return {'status': 'ok'}

    except FileNotFoundError:
        return {
            'status': 'error',
            'type': 'permission',
            'message': 'Failed to obtain launch permission'
        }


def setup_temp_assets():
    """複製網路資源到暫存目錄"""
    try:
        temp_dir = tempfile.gettempdir()
        temp_assets_dir = os.path.join(temp_dir, 'cp_wafer_map_assets')
        
        if not os.path.exists(temp_assets_dir):
            os.makedirs(temp_assets_dir)
        
        fa_source = NETWORK_ASSETS['font_awesome_base']
        fa_dest = os.path.join(temp_assets_dir, 'Font_Awesome')
        if os.path.exists(fa_source) and not os.path.exists(fa_dest):
            shutil.copytree(fa_source, fa_dest)
        
        gf_source = NETWORK_ASSETS['google_fonts_base']
        gf_dest = os.path.join(temp_assets_dir, 'Google_Fonts')
        if os.path.exists(gf_source) and not os.path.exists(gf_dest):
            shutil.copytree(gf_source, gf_dest)
        
        return temp_assets_dir
    except Exception as e:
        print(f"Error setting up temp assets: {str(e)}")
        return None


# ==================== Flask Routes ====================

@app.route('/assets/<path:filename>')
def serve_asset(filename):
    """提供靜態資源檔案"""
    if filename.startswith('Font_Awesome/'):
        base_path = NETWORK_ASSETS['font_awesome_base']
        sub_path = filename.replace('Font_Awesome/', '')
    elif filename.startswith('Google_Fonts/'):
        base_path = NETWORK_ASSETS['google_fonts_base']
        sub_path = filename.replace('Google_Fonts/', '')
    else:
        return "Not found", 404
    
    return send_from_directory(base_path, sub_path)


@app.route('/')
def index():
    """主頁面"""
    update_last_activity()
    return generate_html()


@app.route('/result')
def show_result():
    """顯示分析結果"""
    global analysis_file_data, anomaly_weights
    
    update_last_activity()
    
    if not analysis_file_data:
        return redirect('/')
    
    html_content = create_html_report(analysis_file_data, anomaly_weights)
    return html_content


@app.route('/api/check_version', methods=['GET'])
def api_check_version():
    """版本檢查 API"""
    update_last_activity()
    result = check_version()
    return jsonify(result)


@app.route('/api/execute_update', methods=['POST'])
def execute_update():
    """執行更新：啟動批次檔、開啟安裝程式、退出程式"""
    try:
        update_last_activity()
        
        data = request.get_json()
        latest_exe = data.get('latest_exe')
        
        if not latest_exe:
            return jsonify({'success': False, 'error': 'No latest_exe provided'})
        
        bat_content = '''@echo off
timeout /t 3 /nobreak
rmdir /s /q "C:\\Users\\{username}\\BITools\\Wafer Map Stack Analysis"
del "%~f0"
'''
        bat_path = os.path.join(os.environ['TEMP'], 'delete_Wafer_Map_Stack_Analysis.bat')
        with open(bat_path, 'w') as f:
            f.write(bat_content)
        
        subprocess.Popen(['cmd', '/c', bat_path], shell=True, 
                        creationflags=subprocess.CREATE_NO_WINDOW)
        
        os.startfile(latest_exe)
        
        time.sleep(0.5)
        os._exit(0)
        
        return jsonify({'success': True})
        
    except Exception as e:
        print(f"Error executing update: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/select_folder', methods=['POST'])
def select_folder():
    """選擇資料夾並回傳檔案列表"""
    global selected_folder_path, xlsx_file_list
    
    try:
        update_last_activity()
        
        root = Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        
        folder_path = filedialog.askdirectory(title='Select Folder', parent=root)
        root.destroy()
        
        if not folder_path:
            return jsonify({'success': False, 'error': 'No folder selected'})
        
        selected_folder_path = folder_path
        folder_name = os.path.basename(folder_path)
        
        files = os.listdir(folder_path)
        non_xlsx_files = []
        xlsx_files = []
        temp_files = []
        
        for file in files:
            file_path = os.path.join(folder_path, file)
            if os.path.isfile(file_path):
                if file.startswith('~$'):
                    temp_files.append(file)
                elif not file.endswith('.xlsx'):
                    non_xlsx_files.append(file)
                else:
                    xlsx_files.append(file_path)
        
        if temp_files:
            temp_file_list = '\n'.join(temp_files[:5])
            if len(temp_files) > 5:
                temp_file_list += '\n...'
            
            return jsonify({
                'success': False,
                'error': f'Excel temporary files detected!\n\nFound temporary Excel files (starting with ~$).\nPlease close all Excel files in this folder before proceeding.\n\nTemporary files found:\n{temp_file_list}'
            })
        
        if non_xlsx_files:
            non_xlsx_list = ', '.join(non_xlsx_files[:5])
            if len(non_xlsx_files) > 5:
                non_xlsx_list += '...'
            
            return jsonify({
                'success': False,
                'error': f'Invalid file type detected!\n\nOnly .xlsx files are allowed.\nFound non-xlsx files: {non_xlsx_list}'
            })
        
        if not xlsx_files:
            return jsonify({'success': False, 'error': 'No Excel files found in selected folder'})
        
        xlsx_file_list = xlsx_files
        
        return jsonify({
            'success': True,
            'folder_path': folder_path,
            'folder_name': folder_name,
            'file_count': len(xlsx_files),
            'files': [os.path.basename(f) for f in xlsx_files]
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/start_analysis', methods=['POST'])
def start_analysis():
    """開始分析"""
    global current_worker, analysis_file_data, anomaly_weights, selected_process_type
    
    try:
        update_last_activity()
        
        data = request.get_json()
        process_type = data.get('process_type', 'CP')
        selected_process_type = process_type
        
        if not xlsx_file_list:
            return jsonify({'success': False, 'error': 'Please select a folder first'})
        
        if current_worker and current_worker.is_alive():
            return jsonify({'success': False, 'error': 'Analysis already in progress'})
        
        anomaly_weights = load_anomaly_weights(process_type)
        
        if not anomaly_weights:
            return jsonify({'success': False, 'error': f'Failed to load {process_type} anomaly weights configuration'})
        
        def on_complete(file_data):
            global analysis_file_data
            analysis_file_data = file_data
            print("Analysis completed successfully")
        
        def on_error(title, error_msg):
            print(f"Analysis error [{title}]: {error_msg}")
        
        current_worker = AnalysisWorker(
            xlsx_files=xlsx_file_list,
            anomaly_weights=anomaly_weights,
            complete_callback=on_complete,
            error_callback=on_error
        )
        
        current_worker.start()
        
        return jsonify({'success': True})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/get_progress')
def get_progress():
    """取得分析進度"""
    global current_worker
    
    try:
        update_last_activity()
        
        if not current_worker:
            return jsonify({
                'progress': 0,
                'status': 'idle',
                'message': 'No analysis in progress'
            })
        
        return jsonify({
            'progress': current_worker.current_progress,
            'status': current_worker.status,
            'message': current_worker.status_message,
            'error': current_worker.error_message if current_worker.status == 'error' else ''
        })
        
    except Exception as e:
        return jsonify({
            'progress': 0,
            'status': 'error',
            'message': str(e)
        })


@app.route('/api/heartbeat', methods=['POST'])
def heartbeat():
    """接收前端心跳以保持伺服器活躍"""
    update_last_activity()
    return jsonify({'success': True, 'message': 'Heartbeat received'})


@app.route('/api/shutdown', methods=['POST'])
def shutdown():
    """立即關閉伺服器（當瀏覽器關閉時調用）"""
    print("Browser closed, shutting down immediately...")
    
    def delayed_shutdown():
        time.sleep(0.5)
        os._exit(0)
    
    threading.Thread(target=delayed_shutdown, daemon=True).start()
    return jsonify({'success': True})


@app.route('/api/get_anomaly_codes', methods=['GET'])
def get_anomaly_codes():
    """取得所有異常代碼資訊"""
    global analysis_file_data, anomaly_weights
    
    try:
        update_last_activity()
        
        if not analysis_file_data:
            return jsonify({
                'success': False,
                'error': 'No analysis data available'
            })
        
        codes_data = prepare_anomaly_codes_data(analysis_file_data, anomaly_weights)
        
        return jsonify({
            'success': True,
            'codes': codes_data
        })
        
    except Exception as e:
        print(f"Error in get_anomaly_codes: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        })


@app.route('/api/regenerate_all_charts', methods=['POST'])
def regenerate_all_charts():
    """統一重新生成所有圖表（包含可選的 Mask Overlay）"""
    global analysis_file_data, anomaly_weights
    
    try:
        update_last_activity()
        
        if not analysis_file_data:
            return jsonify({
                'success': False,
                'error': 'No analysis data available'
            })
        
        data = request.get_json()
        selected_codes = data.get('selected_codes', [])
        x_mask_size = data.get('x_mask_size')
        y_mask_size = data.get('y_mask_size')
        
        if not selected_codes:
            return jsonify({
                'success': False,
                'error': 'No codes selected'
            })
        
        # 生成 Distribution Chart
        dist_fig = generate_anomaly_distribution_chart(
            analysis_file_data, 
            anomaly_weights, 
            selected_codes,
            x_mask_size,
            y_mask_size
        )
        
        if dist_fig is None:
            return jsonify({
                'success': False,
                'error': 'Failed to generate distribution chart'
            })
        
        # 生成 Heatmap Chart
        heatmap_fig = generate_anomaly_heatmap_chart(
            analysis_file_data,
            selected_codes,
            x_mask_size,
            y_mask_size
        )
        
        if heatmap_fig is None:
            return jsonify({
                'success': False,
                'error': 'Failed to generate heatmap chart'
            })
        
        # 生成 Mask Overlay Heatmap（如果有 mask 設定）
        mask_fig = None
        if x_mask_size and y_mask_size:
            mask_fig = generate_mask_overlay_heatmap(
                analysis_file_data,
                anomaly_weights,
                selected_codes,
                x_mask_size,
                y_mask_size
            )
        
        # 轉換為字典
        dist_dict = dist_fig.to_dict()
        heatmap_dict = heatmap_fig.to_dict()
        
        result = {
            'success': True,
            'distribution': {
                'data': dist_dict['data'],
                'layout': dist_dict['layout'],
                'config': {
                    'responsive': True,
                    'displayModeBar': True,
                    'toImageButtonOptions': {
                        'format': 'png',
                        'filename': 'anomaly_distribution_filtered'
                    }
                }
            },
            'heatmap': {
                'data': heatmap_dict['data'],
                'layout': heatmap_dict['layout'],
                'config': {
                    'responsive': True,
                    'displayModeBar': True,
                    'toImageButtonOptions': {
                        'format': 'png',
                        'filename': 'anomaly_heatmap_filtered'
                    }
                }
            },
            'mask_heatmap': None
        }
        
        if mask_fig:
            mask_dict = mask_fig.to_dict()
            result['mask_heatmap'] = {
                'data': mask_dict['data'],
                'layout': mask_dict['layout'],
                'config': {
                    'responsive': True,
                    'displayModeBar': True,
                    'toImageButtonOptions': {
                        'format': 'png',
                        'filename': 'mask_overlay_heatmap'
                    }
                }
            }
        
        return jsonify(result)
        
    except Exception as e:
        print(f"Error in regenerate_all_charts: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return jsonify({
            'success': False,
            'error': str(e)
        })

def generate_html():
    """生成主頁面 HTML"""
    html_content = """
<!DOCTYPE html>
<html lang="zh-TW">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Wafer Map Stack Analysis</title>
    <link rel="stylesheet" href="/assets/Google_Fonts/css/noto-sans-tc.css">
    <link rel="stylesheet" href="/assets/Font_Awesome/css/all.min.css">
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }

        body {
            font-family: "Noto Sans TC", Arial, sans-serif;
            background: linear-gradient(135deg, #F5F5F5 0%, #E8E8E8 100%);
            min-height: 100vh;
            display: flex;
            justify-content: center;
            align-items: center;
            padding: 20px;
        }

        .container {
            background: #FFFFFF;
            border-radius: 20px;
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.08), 0 2px 8px rgba(0, 0, 0, 0.04);
            padding: 40px;
            max-width: 600px;
            width: 100%;
            animation: fadeIn 0.5s ease-in;
        }

        @keyframes fadeIn {
            from {
                opacity: 0;
                transform: translateY(20px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }

        .header {
            text-align: center;
            margin-bottom: 40px;
            padding-bottom: 25px;
            border-bottom: 1px solid #E8E8E8;
        }

        .header h1 {
            font-size: 28px;
            color: #2C2C2C;
            margin-bottom: 10px;
            font-weight: 700;
            letter-spacing: -0.5px;
        }

        .header .subtitle {
            font-size: 15px;
            color: #666666;
            font-weight: 400;
        }

        .section {
            margin-bottom: 30px;
        }

        .section-title {
            font-size: 12px;
            color: #666666;
            margin-bottom: 15px;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 1px;
        }

        /* 新增：製程類型選擇器樣式 */
        .process-type-selector {
            display: flex;
            gap: 12px;
            margin-bottom: 20px;
        }

        .process-type-btn {
            flex: 1;
            padding: 16px;
            border: 2px solid #E8E8E8;
            border-radius: 12px;
            background: #FAFAFA;
            color: #666666;
            font-size: 16px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 10px;
            letter-spacing: 0.5px;
            position: relative;
        }

        .process-type-btn:hover {
            border-color: #D0D0D0;
            background: #F5F5F5;
            transform: translateY(-1px);
        }

        .process-type-btn.active {
            background: linear-gradient(135deg, #4A4A4A 0%, #2C2C2C 100%);
            color: #FFFFFF;
            border-color: #2C2C2C;
            box-shadow: 0 4px 12px rgba(44, 44, 44, 0.15);
        }

        .process-type-btn.active:hover {
            box-shadow: 0 6px 16px rgba(44, 44, 44, 0.25);
            transform: translateY(-2px);
        }

        .process-type-btn .icon {
            font-size: 18px;
        }

        .process-type-info {
            background: #F0F8FF;
            border-left: 4px solid #4A4A4A;
            padding: 12px 16px;
            border-radius: 8px;
            margin-top: 12px;
            font-size: 13px;
            color: #555555;
            line-height: 1.6;
        }

        .btn {
            width: 100%;
            padding: 15px;
            border: none;
            border-radius: 12px;
            font-size: 15px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 10px;
            letter-spacing: 0.3px;
        }

        .btn-primary {
            background: linear-gradient(135deg, #4A4A4A 0%, #2C2C2C 100%);
            color: #FFFFFF;
            box-shadow: 0 4px 12px rgba(44, 44, 44, 0.15);
        }

        .btn-primary:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 20px rgba(44, 44, 44, 0.25);
            background: linear-gradient(135deg, #5A5A5A 0%, #3C3C3C 100%);
        }

        .btn-success {
            background: linear-gradient(135deg, #3A3A3A 0%, #1A1A1A 100%);
            color: #FFFFFF;
            box-shadow: 0 4px 12px rgba(26, 26, 26, 0.15);
        }

        .btn-success:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 20px rgba(26, 26, 26, 0.25);
            background: linear-gradient(135deg, #4A4A4A 0%, #2A2A2A 100%);
        }

        .btn:disabled {
            background: #D0D0D0;
            color: #888888;
            cursor: not-allowed;
            transform: none;
            box-shadow: none;
        }

        .file-info {
            background: #FAFAFA;
            border: 1px solid #E8E8E8;
            border-radius: 12px;
            padding: 18px;
            margin-top: 15px;
            display: none;
        }

        .file-info.show {
            display: block;
            animation: slideDown 0.3s ease;
        }

        @keyframes slideDown {
            from {
                opacity: 0;
                max-height: 0;
            }
            to {
                opacity: 1;
                max-height: 200px;
            }
        }

        .file-info-item {
            display: flex;
            justify-content: space-between;
            padding: 10px 0;
            border-bottom: 1px solid #E8E8E8;
        }

        .file-info-item:last-child {
            border-bottom: none;
        }

        .file-info-label {
            font-weight: 600;
            color: #666666;
            font-size: 14px;
        }

        .file-info-value {
            color: #2C2C2C;
            font-size: 14px;
            font-weight: 500;
        }

        .progress-container {
            margin-top: 20px;
            display: none;
        }

        .progress-container.show {
            display: block;
            animation: fadeIn 0.3s ease;
        }

        .progress-bar-wrapper {
            width: 100%;
            height: 32px;
            background: #E8E8E8;
            border-radius: 16px;
            overflow: hidden;
            position: relative;
            box-shadow: inset 0 2px 4px rgba(0, 0, 0, 0.06);
        }

        .progress-bar {
            height: 100%;
            background: linear-gradient(90deg, #5A5A5A 0%, #3A3A3A 100%);
            width: 0%;
            transition: width 0.3s ease;
            display: flex;
            align-items: center;
            justify-content: center;
            color: white;
            font-weight: 600;
            font-size: 13px;
            letter-spacing: 0.3px;
        }

        .status-text {
            text-align: center;
            margin-top: 12px;
            color: #666666;
            font-size: 13px;
            font-weight: 500;
        }

        .modal {
            display: none;
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background: rgba(0, 0, 0, 0.45);
            backdrop-filter: blur(4px);
            justify-content: center;
            align-items: center;
            z-index: 1000;
        }

        .modal.show {
            display: flex;
            animation: fadeIn 0.3s ease;
        }

        .modal-content {
            background: white;
            border-radius: 16px;
            padding: 32px;
            max-width: 400px;
            width: 90%;
            text-align: center;
            animation: slideUp 0.3s ease;
            box-shadow: 0 12px 48px rgba(0, 0, 0, 0.15);
        }

        @keyframes slideUp {
            from {
                opacity: 0;
                transform: translateY(50px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }

        .modal-icon {
            font-size: 48px;
            margin-bottom: 20px;
            color: #4A4A4A;
        }

        .modal-icon.error {
            color: #E74C3C;
        }

        .modal-title {
            font-size: 20px;
            font-weight: 700;
            margin-bottom: 12px;
            color: #2C2C2C;
        }

        .modal-message {
            font-size: 14px;
            color: #666666;
            margin-bottom: 24px;
            white-space: pre-line;
            line-height: 1.6;
        }

        .modal-buttons {
            display: flex;
            gap: 10px;
            justify-content: center;
        }

        .modal-btn {
            padding: 12px 24px;
            border: none;
            border-radius: 10px;
            font-size: 14px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s ease;
            background: linear-gradient(135deg, #4A4A4A 0%, #2C2C2C 100%);
            color: white;
            letter-spacing: 0.3px;
            flex: 1;
        }

        .modal-btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 16px rgba(44, 44, 44, 0.25);
        }

        .modal-btn-secondary {
            background: #E8E8E8;
            color: #2D2D2D;
        }

        .modal-btn-secondary:hover {
            background: #D8D8D8;
            box-shadow: 0 6px 16px rgba(0, 0, 0, 0.1);
        }

        .icon-spin {
            animation: spin 1s linear infinite;
        }

        @keyframes spin {
            from {
                transform: rotate(0deg);
            }
            to {
                transform: rotate(360deg);
            }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Wafer Map Stack Analysis</h1>
            <p class="subtitle">Data Visualization Analytics Tool</p>
        </div>

        <!-- 新增：製程類型選擇區塊 -->
        <div class="section">
            <div class="section-title">Step 1: Select Process Type</div>
            <div class="process-type-selector">
                <button class="process-type-btn active" data-type="CP">
                    <span class="icon"><i class="fas fa-microchip"></i></span>
                    <span>CP</span>
                </button>
                <button class="process-type-btn" data-type="PSI">
                    <span class="icon"><i class="fas fa-layer-group"></i></span>
                    <span>PSI</span>
                </button>
            </div>
            <div class="process-type-info">
                <strong>Current Selection:</strong> <span id="selectedProcessType">CP (CP Process)</span>
            </div>
        </div>

        <div class="section">
            <div class="section-title">Step 2: Select Data Folder</div>
            <button class="btn btn-primary" id="selectFolderBtn">
                <i class="fas fa-folder-open"></i>
                Select Folder
            </button>
            <div class="file-info" id="fileInfo">
                <div class="file-info-item">
                    <span class="file-info-label">Folder Name:</span>
                    <span class="file-info-value" id="folderName">-</span>
                </div>
                <div class="file-info-item">
                    <span class="file-info-label">Excel Files Found:</span>
                    <span class="file-info-value" id="fileCount">0</span>
                </div>
            </div>
        </div>

        <div class="section">
            <div class="section-title">Step 3: Start Analysis</div>
            <button class="btn btn-success" id="startAnalysisBtn" disabled>
                <i class="fas fa-play"></i>
                Start Analysis
            </button>
        </div>

        <div class="progress-container" id="progressContainer">
            <div class="progress-bar-wrapper">
                <div class="progress-bar" id="progressBar">0%</div>
            </div>
            <div class="status-text" id="statusText">Initializing...</div>
        </div>
    </div>

    <div class="modal" id="errorModal">
        <div class="modal-content">
            <div class="modal-icon error">
                <i class="fas fa-exclamation-circle"></i>
            </div>
            <div class="modal-title">Error</div>
            <div class="modal-message" id="errorMessage">
                An error occurred during analysis.
            </div>
            <div class="modal-buttons">
                <button class="modal-btn" id="closeErrorBtn">Close</button>
            </div>
        </div>
    </div>

    <div class="modal" id="messageModal">
        <div class="modal-content">
            <div class="modal-icon" id="messageIcon">
                <i class="fa-solid fa-circle-info"></i>
            </div>
            <div class="modal-title" id="messageTitle">Message</div>
            <div class="modal-message" id="messageText">Message content</div>
            <div class="modal-buttons">
                <button class="modal-btn" id="messageOkBtn">OK</button>
            </div>
        </div>
    </div>

    <div class="modal" id="reminderModal">
        <div class="modal-content">
            <div class="modal-icon">
                <i class="fas fa-info-circle"></i>
            </div>
            <div class="modal-title">Reminder</div>
            <div class="modal-message">
                請確保資料夾內存放的檔案都屬同一個 Group!!
            </div>
            <div class="modal-buttons">
                <button class="modal-btn" id="reminderOkBtn">確定</button>
            </div>
        </div>
    </div>

    <div class="modal" id="confirmModal">
        <div class="modal-content">
            <div class="modal-icon">
                <i class="fas fa-question-circle"></i>
            </div>
            <div class="modal-title" id="confirmTitle">Confirm</div>
            <div class="modal-message" id="confirmMessage">Message</div>
            <div class="modal-buttons">
                <button class="modal-btn modal-btn-secondary" id="confirmCancelBtn">Cancel</button>
                <button class="modal-btn" id="confirmOkBtn">Confirm</button>
            </div>
        </div>
    </div>

    <script>
        let isNormalNavigation = false;
        
        window.addEventListener('beforeunload', () => {
            if (!isNormalNavigation) {
                fetch('/api/shutdown', {
                    method: 'POST',
                    keepalive: true
                });
            }
        });
        
        setInterval(() => {
            fetch('/api/heartbeat', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' }
            }).catch(error => {
                console.error('Heartbeat failed:', error);
            });
        }, 60000);
        
        // 新增：製程類型選擇邏輯
        let selectedProcessType = 'CP';
        
        const processTypeBtns = document.querySelectorAll('.process-type-btn');
        const selectedProcessTypeText = document.getElementById('selectedProcessType');
        
        processTypeBtns.forEach(btn => {
            btn.addEventListener('click', () => {
                // 移除所有按鈕的 active 狀態
                processTypeBtns.forEach(b => b.classList.remove('active'));
                
                // 添加當前按鈕的 active 狀態
                btn.classList.add('active');
                
                // 更新選中的製程類型
                selectedProcessType = btn.dataset.type;
                
                // 更新顯示文字
                const displayText = selectedProcessType === 'CP' 
                    ? 'CP (CP Process)' 
                    : 'PSI (WLCSP Process)';
                selectedProcessTypeText.textContent = displayText;
                
                console.log('Selected process type:', selectedProcessType);
            });
        });
        
        const selectFolderBtn = document.getElementById('selectFolderBtn');
        const startAnalysisBtn = document.getElementById('startAnalysisBtn');
        const fileInfo = document.getElementById('fileInfo');
        const folderName = document.getElementById('folderName');
        const fileCount = document.getElementById('fileCount');
        const progressContainer = document.getElementById('progressContainer');
        const progressBar = document.getElementById('progressBar');
        const statusText = document.getElementById('statusText');
        const errorModal = document.getElementById('errorModal');
        const errorMessage = document.getElementById('errorMessage');
        const closeErrorBtn = document.getElementById('closeErrorBtn');
        const messageModal = document.getElementById('messageModal');
        const messageIcon = document.getElementById('messageIcon');
        const messageTitle = document.getElementById('messageTitle');
        const messageText = document.getElementById('messageText');
        const messageOkBtn = document.getElementById('messageOkBtn');
        const reminderModal = document.getElementById('reminderModal');
        const reminderOkBtn = document.getElementById('reminderOkBtn');

        let analysisInProgress = false;

        async function showConfirm(message, isAlert = false) {
            return new Promise((resolve) => {
                const confirmModal = document.getElementById('confirmModal');
                const confirmMessage = document.getElementById('confirmMessage');
                const confirmOkBtn = document.getElementById('confirmOkBtn');
                const confirmCancelBtn = document.getElementById('confirmCancelBtn');
                
                confirmMessage.textContent = message;
                
                if (isAlert) {
                    confirmCancelBtn.style.display = 'none';
                    confirmOkBtn.textContent = 'OK';
                } else {
                    confirmCancelBtn.style.display = 'block';
                    confirmOkBtn.textContent = 'Confirm';
                }
                
                confirmModal.classList.add('show');
                
                confirmOkBtn.onclick = () => {
                    confirmModal.classList.remove('show');
                    resolve(true);
                };
                
                confirmCancelBtn.onclick = () => {
                    confirmModal.classList.remove('show');
                    resolve(false);
                };
            });
        }

        async function checkVersionOnStartup() {
            try {
                const response = await fetch('/api/check_version');
                const result = await response.json();

                if (result.status === 'update') {
                    messageIcon.innerHTML = '<i class="fa-solid fa-circle-info"></i>';
                    messageIcon.style.color = '#4A4A4A';
                    messageTitle.textContent = 'Version Update Available';
                    messageText.textContent = result.message;
                    messageModal.style.display = 'flex';

                    messageOkBtn.onclick = async () => {
                        messageModal.style.display = 'none';

                        try {
                            await fetch('/api/execute_update', {
                                method: 'POST',
                                headers: { 'Content-Type': 'application/json' },
                                body: JSON.stringify({ latest_exe: result.latest_exe }),
                                keepalive: true
                            });
                        } catch (error) {
                            console.error('Failed to exit:', error);
                        }

                        window.close();
                        window.location.href = 'about:blank';
                    };
                } else if (result.status === 'error') {
                    showError(result.message);
                    if (result.type === 'permission' || result.type === 'version') {
                        document.querySelectorAll('button, select, input').forEach(el => {
                            el.disabled = true;
                        });
                    }
                }
            } catch (error) {
                console.error('Version check failed:', error);
            }
        }

        window.addEventListener('DOMContentLoaded', checkVersionOnStartup);

        selectFolderBtn.addEventListener('click', () => {
            reminderModal.classList.add('show');
        });

        reminderOkBtn.addEventListener('click', async () => {
            reminderModal.classList.remove('show');
            
            try {
                selectFolderBtn.disabled = true;
                selectFolderBtn.innerHTML = '<i class="fas fa-spinner icon-spin"></i> Loading...';

                const response = await fetch('/api/select_folder', {
                    method: 'POST'
                });
                const result = await response.json();

                if (result.success) {
                    folderName.textContent = result.folder_name;
                    fileCount.textContent = result.file_count;
                    fileInfo.classList.add('show');
                    startAnalysisBtn.disabled = false;
                } else {
                    showError(result.error);
                }
            } catch (error) {
                showError('Failed to select folder: ' + error);
            } finally {
                selectFolderBtn.disabled = false;
                selectFolderBtn.innerHTML = '<i class="fas fa-folder-open"></i> Select Folder';
            }
        });

        startAnalysisBtn.addEventListener('click', async () => {
            try {
                startAnalysisBtn.disabled = true;
                selectFolderBtn.disabled = true;
                progressContainer.classList.add('show');
                analysisInProgress = true;

                // 修改：將 selectedProcessType 傳送到後端
                const response = await fetch('/api/start_analysis', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json'
                    },
                    body: JSON.stringify({
                        process_type: selectedProcessType
                    })
                });
                const result = await response.json();

                if (result.success) {
                    pollProgress();
                } else {
                    showError(result.error);
                    resetUI();
                }
            } catch (error) {
                showError('Failed to start analysis: ' + error);
                resetUI();
            }
        });

        function pollProgress() {
            const interval = setInterval(async () => {
                try {
                    const response = await fetch('/api/get_progress');
                    const progress = await response.json();

                    progressBar.style.width = progress.progress + '%';
                    progressBar.textContent = progress.progress + '%';
                    statusText.textContent = progress.message;

                    if (progress.status === 'completed') {
                        clearInterval(interval);
                        analysisInProgress = false;
                        isNormalNavigation = true;
                        window.location.href = '/result';
                    } else if (progress.status === 'error') {
                        clearInterval(interval);
                        analysisInProgress = false;
                        showError(progress.error || progress.message);
                        resetUI();
                    }
                } catch (error) {
                    clearInterval(interval);
                    analysisInProgress = false;
                    showError('Failed to check progress: ' + error);
                    resetUI();
                }
            }, 500);
        }

        closeErrorBtn.addEventListener('click', () => {
            errorModal.classList.remove('show');
        });

        function showError(message) {
            errorMessage.textContent = message;
            errorModal.classList.add('show');
        }

        function resetUI() {
            startAnalysisBtn.disabled = false;
            selectFolderBtn.disabled = false;
            progressContainer.classList.remove('show');
            progressBar.style.width = '0%';
            progressBar.textContent = '0%';
            statusText.textContent = 'Initializing...';
        }

        window.addEventListener('click', (e) => {
            if (e.target === errorModal) {
                errorModal.classList.remove('show');
            }
            if (e.target === messageModal) {
                messageModal.classList.remove('show');
            }
            if (e.target === reminderModal) {
                reminderModal.classList.remove('show');
            }
            if (e.target === document.getElementById('confirmModal')) {
                document.getElementById('confirmModal').classList.remove('show');
            }
        });
    </script>
</body>
</html>
    """
    return html_content

# ==================== 主程式進入點 ====================

def main():
    """主程式啟動函數"""
    save_log()
    
    version_result = check_version()
    if version_result['status'] == 'error' and version_result.get('type') == 'permission':
        print(version_result['message'])
        sys.exit(1)
    
    try:
        port = find_available_port()
        print(f"Starting server on port {port}...")
    except RuntimeError as e:
        print(f"Error: {e}")
        return
    
    activity_thread = threading.Thread(target=check_activity_thread, daemon=True)
    activity_thread.start()
    
    url = f'http://localhost:{port}'
    threading.Timer(1.0, lambda: webbrowser.open(url)).start()
    
    app.run(host='localhost', port=port, debug=False, threaded=True)


if __name__ == '__main__':
    main()